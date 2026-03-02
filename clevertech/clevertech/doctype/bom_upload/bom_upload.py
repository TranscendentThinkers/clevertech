import frappe
from frappe.model.document import Document
from frappe.utils.file_manager import save_file
import openpyxl
import io
import hashlib
import json

# Conditional import for image loader
try:
    from openpyxl_image_loader import SheetImageLoader
    HAS_IMAGE_LOADER = True
except ImportError:
    HAS_IMAGE_LOADER = False
    frappe.log_error("openpyxl_image_loader not installed. Image upload will be skipped.", "BOM Upload")


class BOMUpload(Document):
    pass


@frappe.whitelist()
def create_boms(docname):
    doc = frappe.get_doc("BOM Upload", docname)
    return _create_boms(doc)


def _create_boms(doc):

    file_doc = frappe.get_doc("File", {"file_url": doc.bom_file})
    content = file_doc.get_content()

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Initialize image loader only if available
    image_loader = SheetImageLoader(ws) if HAS_IMAGE_LOADER else None

    rows = parse_rows(ws)
    tree = build_tree(rows)
    created = 0
    skipped = 0
    failed = 0
    errors = []

    for node in tree:
        try:
            created_flag = create_bom_recursive(node, doc.project, ws, image_loader)
            if created_flag:
                created += 1
            else:
                skipped += 1
        except Exception as e:
            failed += 1
            errors.append(f"{node['item_code']} → {str(e)}")

    return {
        "created": created,
        "skipped": skipped,
        "failed": failed,
        "errors": errors
    }


# ================= Excel helpers =================

def clean_code(code):
    if not code:
        return None
    return str(code).replace(".", "").strip()


def to_float(val, default=0):
    try:
        return float(val)
    except Exception:
        return default
def _calculate_tree_hash(children):
    """
    Calculate MD5 hash of BOM structure from tree node children.
    Uses same logic as bom_hooks.calculate_bom_structure_hash for consistency.

    Args:
        children: List of child node dicts with item_code and qty

    Returns:
        str: MD5 hash of sorted (item_code, qty) tuples, or None if no children
    """
    if not children:
        return None

    structure = sorted(
        [(child["item_code"], float(child.get("qty", 1))) for child in children],
        key=lambda x: x[0],
    )
    structure_str = json.dumps(structure, sort_keys=True)
    return hashlib.md5(structure_str.encode()).hexdigest()


def normalize_uom(uom):
    if not uom:
        return "Nos"

    # Convert to string and clean
    u = str(uom).strip().upper()

    # If it's a number or doesn't match known UOMs, default to Nos
    if u.isdigit():
        return "Nos"

    mapping = {
        "NUMERI": "Nos",
        "PEZZI": "PIECES",
        "METRI": "Meter",
        "MQ": "Square Meter",  # Metro Quadro (Square Meter)
        "PACKAGES": "Nos"
    }

    # Return mapped value, or default to Nos if not found
    return mapping.get(u, "Nos")


# ================= Database-Driven Mapping Functions =================

def get_item_group_and_hsn(item_code):
    """Map item code prefix to item group, HSN code, and expense account using Item Denomination Map"""
    if not item_code:
        return "All Item Groups", None, None

    code = str(item_code).upper()

    # Check for multi-character prefixes first (longer matches first)
    prefixes_to_check = ["CIM", "CIG", "CIC", "CIE", "IM", "TS"]
    for prefix in prefixes_to_check:
        if code.startswith(prefix):
            mapping = frappe.db.get_value(
                "Item Denomination Map",
                {"denomination": prefix},
                ["item_group", "hsn_code", "default_expense_account"],
                as_dict=True
            )
            if mapping:
                return mapping.item_group, mapping.hsn_code, mapping.default_expense_account

    # Check single character prefixes
    first_char = code[0] if code else None
    if first_char:
        mapping = frappe.db.get_value(
            "Item Denomination Map",
            {"denomination": first_char},
            ["item_group", "hsn_code", "default_expense_account"],
            as_dict=True
        )
        if mapping:
            return mapping.item_group, mapping.hsn_code, mapping.default_expense_account

    return "All Item Groups", None, None


def normalize_material(material):
    """Map material values to standardized names using Material Mapping doctype"""
    if not material:
        return None

    mat = str(material).strip()

    # Try to get mapping from database
    mapped_material = frappe.db.get_value(
        "Material Mapping",
        {"materiale": mat},
        "material"
    )

    # If found in database, return it; otherwise return original
    return mapped_material if mapped_material else mat


def get_type_of_material(description):
    """Extract first word from description and map to type of material using Type of Material doctype"""
    if not description:
        return None

    first_word = str(description).strip().split()[0].upper()

    # Try to get mapping from database
    type_of_mat = frappe.db.get_value(
        "Type of Material",
        {"item_description": first_word},
        "type_of_material"
    )

    return type_of_mat
def get_surface_treatment(treatment):
    """Map Italian surface treatment to English using Surface Treatment Translation doctype"""
    if not treatment:
        return None

    treat = str(treatment).strip()

    # Try to get translation from database
    translated = frappe.db.get_value(
        "Surface Treatment Translation",
        {"italian": treat},
        "english"
    )

    # If found in database, return it; otherwise return original
    return translated if translated else treat


# ================= Parse Excel =================

def parse_rows(ws):

    rows = []

    for r in range(3, ws.max_row + 1):

        code_raw = ws[f"C{r}"].value
        if not code_raw:
            continue

        position = ws[f"A{r}"].value or ws[f"B{r}"].value

        rows.append({
            "row_num": r,  # Store row number for image extraction
            "position": position,
            "item_code": clean_code(code_raw),
            "description": ws[f"D{r}"].value,
            "extended_description": ws[f"U{r}"].value,  # DESCRIZIONE_ESTESA for item description
            "qty": to_float(ws[f"E{r}"].value, 1),
            "revision": ws[f"G{r}"].value,
            "state": ws[f"J{r}"].value,  # State from column J
            "material": ws[f"AD{r}"].value,
            "part_number": ws[f"AE{r}"].value,
            "weight": to_float(ws[f"AF{r}"].value, 0),
            "manufacturer": ws[f"AG{r}"].value,
            "treatment": ws[f"AL{r}"].value,
            "uom": normalize_uom(ws[f"AN{r}"].value),
            "level": int(ws[f"AR{r}"].value or 0),
            "children": []
        })

    return rows


# ================= Hierarchy =================

def build_tree(rows):

    stack = []
    roots = []

    for row in rows:

        while stack and stack[-1]["level"] >= row["level"]:
            stack.pop()
        if stack:
            stack[-1]["children"].append(row)
        else:
            roots.append(row)

        stack.append(row)

    return roots


# ================= Items =================
def ensure_item_exists(item_code, description, extended_description, uom, row_num, ws, image_loader, material=None, treatment=None, weight=None, part_number=None, manufacturer=None, revision=None):
    """
    Create Item master if it doesn't exist.
    If item exists, update fields that have changed.

    Returns:
        str: "created" | "existing" | "updated" | "failed"
    """
    if frappe.db.exists("Item", item_code):
        # Item exists - check if we need to update any fields
        existing_item = frappe.get_doc("Item", item_code)
        
        # Check if revision has changed - only update if revision is different
        existing_revision = existing_item.get("custom_revision_no")
        new_revision = revision if revision is not None and revision != "" else None
        
        # If revisions match, skip update
        if existing_revision == new_revision:
            return "existing"
        
        # Revision has changed, proceed with updates
        updated = False

        # Prepare new values
        item_name = description or item_code
        item_description = extended_description or description or item_code
        item_group, hsn_code, default_expense_account = get_item_group_and_hsn(item_code)
        normalized_material = normalize_material(material)
        type_of_material = get_type_of_material(description)
        surface_treatment = get_surface_treatment(treatment)

        # Check and update basic fields
        if existing_item.item_name != item_name:
            existing_item.item_name = item_name
            updated = True
        
        if existing_item.description != item_description:
            existing_item.description = item_description
            updated = True
        
        if existing_item.item_group != item_group:
            existing_item.item_group = item_group
            updated = True
        
        if uom and existing_item.stock_uom != uom:
            existing_item.stock_uom = uom
            updated = True

        # Check and update custom fields
        if hsn_code and existing_item.get("gst_hsn_code") != hsn_code:
            existing_item.gst_hsn_code = hsn_code
            updated = True
        
        if normalized_material and existing_item.get("custom_material") != normalized_material:
            existing_item.custom_material = normalized_material
            updated = True
        
        if type_of_material and existing_item.get("custom_type_of_material") != type_of_material:
            existing_item.custom_type_of_material = type_of_material
            updated = True
        
        if surface_treatment and existing_item.get("custom_class_name") != surface_treatment:
            existing_item.custom_class_name = surface_treatment
            updated = True
        
        if weight and existing_item.get("custom_last_updating_of") != weight:
            existing_item.custom_last_updating_of = weight
            updated = True
        
        if part_number and existing_item.get("custom_excode") != part_number:
            existing_item.custom_excode = part_number
            updated = True
        
        if manufacturer and existing_item.get("custom_item_short_description") != manufacturer:
            existing_item.custom_item_short_description = manufacturer
            updated = True
        
        # Update revision number (we know it's changed at this point)
        if new_revision is not None:
            existing_item.custom_revision_no = new_revision
            updated = True

        # Update item defaults for expense account if needed
        if default_expense_account:
            existing_defaults = [d for d in existing_item.get("item_defaults", []) 
                               if d.company == "Clevertech Packaging Automation Solutions Pvt. Ltd."]
            
            if not existing_defaults:
                existing_item.append("item_defaults", {
                    "company": "Clevertech Packaging Automation Solutions Pvt. Ltd.",
                    "expense_account": default_expense_account
                })
                updated = True
            elif existing_defaults[0].expense_account != default_expense_account:
                existing_defaults[0].expense_account = default_expense_account
                updated = True

        # Save if any updates were made
        if updated:
            try:
                existing_item.flags.ignore_validate = True
                existing_item.save(ignore_permissions=True)
                frappe.log_error(f"Item updated: {item_code}", "BOM Upload Item Update")
                return "updated"
            except Exception as e:
                frappe.log_error(f"Item update failed for {item_code}: {str(e)}", "BOM Upload Item Error")
                return "failed"

        return "existing"

    item_name = description or item_code
    item_description = extended_description or description or item_code

    # Get item group, HSN code, and expense account based on item code prefix
    item_group, hsn_code, default_expense_account = get_item_group_and_hsn(item_code)

    # Normalize material
    normalized_material = normalize_material(material)

    # Get type of material from description
    type_of_material = get_type_of_material(description)

    # Get surface treatment in English
    surface_treatment = get_surface_treatment(treatment)

    item = frappe.get_doc({
        "doctype": "Item",
        "item_code": item_code,
        "item_name": item_name,
        "description": item_description,
        "item_group": item_group,
        "stock_uom": uom or "Nos",
        "is_stock_item": 1,
        "is_purchase_item": 1,
        "is_sales_item": 1
    })

    # Add custom fields
    if hsn_code:
        item.gst_hsn_code = hsn_code
    if normalized_material:
        item.custom_material = normalized_material
    if type_of_material:
        item.custom_type_of_material = type_of_material
    if surface_treatment:
        item.custom_class_name = surface_treatment

    # Map weight, part_number, and manufacturer to custom fields
    if weight:
        item.custom_last_updating_of = weight
    if part_number:
        item.custom_excode = part_number
    if manufacturer:
        item.custom_item_short_description = manufacturer
    if revision is not None and revision != "":
        item.custom_revision_no = revision

    # Add default expense account to item defaults
    if default_expense_account:
        item.append("item_defaults", {
            "company": "Clevertech Packaging Automation Solutions Pvt. Ltd.",
            "expense_account": default_expense_account
        })

    # Add tax templates (hardcoded like the macro)
    item.append("taxes", {
        "item_tax_template": "GST 18%",
        "tax_category": "In-State"
    })
    item.append("taxes", {
        "item_tax_template": "GST 18%",
        "tax_category": "Out-State"
    })

    # Handle image upload (only if image_loader is available)
    if HAS_IMAGE_LOADER and image_loader:
        try:
            image_cell = f"A{row_num}"
            if image_loader.image_in(image_cell):
                img = image_loader.get(image_cell)
                img_bytes = io.BytesIO()
                img.save(img_bytes, format="PNG")

                file_obj = save_file(
                    f"{item_code}.png",
                    img_bytes.getvalue(),
                    "Item",
                    item_code,
                    is_private=0
                )

                item.image = file_obj.file_url
        except Exception as e:
            # Silently skip image upload if it fails
            frappe.log_error(f"Image upload failed for {item_code}: {str(e)}", "BOM Upload Image Error")

    try:
        item.insert(ignore_permissions=True)
        return "created"
    except Exception as e:
        frappe.log_error(f"Item creation failed for {item_code}: {str(e)}", "BOM Upload Item Error")
        return "failed"


# ================= BOM =================
def create_bom_recursive(node, project, ws, image_loader, root_level=None):
    """Create BOMs recursively, creating child BOMs first (bottom-up)"""

    item_code = node["item_code"]
    debug_info = [f"=== create_bom_recursive: {item_code} ==="]

    if root_level is None:
        root_level = node["level"]

    ensure_item_exists(
        item_code,
        node["description"],
        node.get("extended_description"),
        node["uom"],
        node["row_num"],
        ws,
        image_loader,
        node.get("material"),
        node.get("treatment"),
        node.get("weight"),
        node.get("part_number"),
        node.get("manufacturer"),
        node.get("revision")
    )

    # First, recursively create BOMs for all children that have sub-assemblies
    for child in node["children"]:
        if child["children"]:  # If child has its own children, it needs a BOM
            create_bom_recursive(child, project, ws, image_loader, child["level"])

    # Only create BOM if this item has children (raw materials)
    if not node["children"]:
        return False

    # Hash-based existence check (2026-02-05):
    # - If no active default BOM exists → create new
    # - If BOM exists with SAME hash → skip (unchanged)
    # - If BOM exists with DIFFERENT hash → create new version (ERPNext auto-demotes old one)
    existing_bom = frappe.db.get_value(
        "BOM",
        {
            "item": item_code,
            "is_active": 1,
            "is_default": 1,
            "docstatus": 1
        },
        ["name", "custom_bom_structure_hash"],
        as_dict=True
    )

    new_hash = _calculate_tree_hash(node["children"])
    debug_info.append(f"children_count: {len(node['children'])}")
    debug_info.append(f"new_hash: {new_hash}")

    if existing_bom:
        debug_info.append(f"existing_bom: {existing_bom.name}")
        debug_info.append(f"existing_hash: {existing_bom.custom_bom_structure_hash}")
        debug_info.append(f"hash_match: {existing_bom.custom_bom_structure_hash == new_hash}")

        if existing_bom.custom_bom_structure_hash == new_hash:
            # Same structure, skip creation
            debug_info.append("ACTION: SKIP (hash match)")
            frappe.log_error(title=f"DEBUG: BOM Skip - {item_code}", message="\n".join(debug_info))
            return False

        debug_info.append("ACTION: CREATE NEW VERSION (hash mismatch)")
    else:
        debug_info.append("existing_bom: None")
        debug_info.append("ACTION: CREATE NEW BOM")

    frappe.log_error(title=f"DEBUG: BOM Create - {item_code}", message="\n".join(debug_info))

    # Create BOM for current node (new or version change)
    bom = frappe.get_doc({
        "doctype": "BOM",
        "item": node["item_code"],
        "quantity": 1,
        "project": project,
        "is_active": 1,
        "is_default": 1
    })

    # Add State field to BOM if present
    if node.get("state"):
        bom.custom_state = node["state"]

    # Add ONLY direct children (not grandchildren)
    for child in node["children"]:
        ensure_item_exists(
            child["item_code"],
            child["description"],
            child.get("extended_description"),
            child["uom"],
            child["row_num"],
            ws,
            image_loader,
            child.get("material"),
            child.get("treatment"),
            child.get("weight"),
            child.get("part_number"),
            child.get("manufacturer"),
            child.get("revision")
        )
        bom_item = {
            "item_code": child["item_code"],
            "qty": child["qty"],
            "uom": child["uom"],
            "custom_position": child["position"],
            "custom_revision_no": child["revision"],
            "custom_material": child["material"],
            "custom_part_number": child["part_number"],
            "custom_weight": child["weight"],
            "custom_manufacturer": child["manufacturer"],
            "custom_tipo_trattamento": child["treatment"],
            "custom_level_of_bom": 1  # Always 1 for direct children
        }

        # If child has its own BOM, link to it
        if child["children"]:
            child_bom = frappe.db.get_value("BOM", {
                "item": child["item_code"],
                "is_active": 1,
                "is_default": 1
            }, "name")
            if child_bom:
                bom_item["bom_no"] = child_bom

        bom.append("items", bom_item)

    bom.insert(ignore_permissions=True)
    bom.submit()
    # Log success with BOM name
    frappe.log_error(
        title=f"DEBUG: BOM Created - {item_code}",
        message=f"BOM: {bom.name} | Items: {len(bom.items)} | State: {node.get('state')} | Project: {project}"
    )

    return True
