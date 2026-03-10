"""
BOM Upload - Phase 1

Simplified BOM upload: Items + BOMs + Component Masters.
No MR/PO/RFQ blocking. No diff details. No remarks required.

If existing BOMs will be versioned (hash mismatch), shows a simple
confirm dialog. User clicks Yes → proceed, No → cancel cleanly.

Sequence:
  1. Parse Excel (dynamic column mapping)
  2. Filter non-released G-codes
  3. Create Items for all nodes
  4. Loose item check (block if is_loose_item=1 and can_be_converted_to_bom=0)
  5. Create Component Masters (with Make/Buy prefix logic)
  6. Scan tree for BOM version changes → return needs_confirmation if any found
  7. Create BOMs bottom-up (hash-based: skip if unchanged, new version if changed)
  8. Link active_bom to CMs, populate hierarchy codes, recalculate quantities
  9. Log upload_history
"""

import frappe
from frappe import _
import openpyxl
import io

from clevertech.clevertech.doctype.bom_upload.bom_upload import (
    build_tree,
    create_bom_recursive,
    _calculate_tree_hash,
    HAS_IMAGE_LOADER,
)

from clevertech.clevertech.doctype.bom_upload.bom_upload_enhanced import (
    parse_rows_dynamic,
    filter_tree_by_g_code_state,
    collect_state_warnings,
    ensure_items_for_all_nodes,
    create_component_masters_for_all_items,
    _link_boms_to_component_masters,
    _populate_hierarchy_codes,
    _refresh_bom_usage_hierarchy_codes,
    _get_all_nodes,
)

try:
    from openpyxl_image_loader import SheetImageLoader
except ImportError:
    SheetImageLoader = None


# ==================== Main Entry Point ====================

@frappe.whitelist()
def create_boms_phase1(docname, confirmed=False, state_confirmed=False):
    """
    Phase 1 BOM Upload.

    Args:
        docname:         BOM Upload document name
        confirmed:       If True, skip version-change confirmation and proceed directly
        state_confirmed: If True, skip state-warning confirmation and proceed directly

    Returns:
        dict with status:
          "needs_state_confirmation" → non-Released items found, JS shows confirm dialog
          "needs_confirmation"       → version changes detected, JS shows confirm dialog
          "success"                  → upload complete, JS shows summary
    """
    confirmed = frappe.utils.cint(confirmed)
    state_confirmed = frappe.utils.cint(state_confirmed)

    doc = frappe.get_doc("BOM Upload", docname)

    if not doc.bom_file:
        frappe.throw(_("Please attach a BOM Excel file first."))
    if not doc.project:
        frappe.throw(_("Please select a Project first."))
    if not doc.machine_code:
        frappe.throw(_("Machine Code is required. Please enter the machine code (e.g., P0000000003033)."))

    machine_code = doc.machine_code

    # Step 1: Parse Excel
    file_doc = frappe.get_doc("File", {"file_url": doc.bom_file})
    content = file_doc.get_content()

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Validate root item code from row 1 header (e.g. "Item no:V00000000016")
    root_item_code = None
    for col_num in range(1, 20):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        cell_val = str(ws[f"{col_letter}1"].value or "")
        if "Item no:" in cell_val:
            root_item_code = cell_val.split("Item no:")[-1].strip()
            break
    if root_item_code and not root_item_code.upper().startswith(("P", "V")):
        frappe.throw(
            _(f"Kindly upload Valid Machine Code BOM file. "
              f"The Item should start with either P or V. Found: {root_item_code}"),
            title=_("Invalid BOM File")
        )

    image_loader = None
    if HAS_IMAGE_LOADER and SheetImageLoader:
        try:
            image_loader = SheetImageLoader(ws)
        except Exception:
            pass

    rows = parse_rows_dynamic(ws)
    tree = build_tree(rows)

    if not tree:
        frappe.throw(_("No components found in the Excel file."))

    # Step 2: Filter non-released G-codes
    filtered_tree, skipped_info = filter_tree_by_g_code_state(tree)

    if skipped_info["count"] > 0:
        skip_details = "<br>".join([
            f"• <b>{item['item_code']}</b> — STATE: {item['state'] or '(blank)'}"
            for item in skipped_info["items"]
        ])
        frappe.msgprint(
            f"<b>Skipped {skipped_info['count']} non-released G-code(s)</b><br><br>"
            f"Total items skipped (including children): <b>{skipped_info['total_items_skipped']}</b><br><br>"
            f"{skip_details}",
            title=_("Non-Released G-Codes Skipped"),
            indicator="orange"
        )

    tree = filtered_tree

    if not tree:
        frappe.throw(_("No releasable components found after G-code STATE filtering."))

    # Step 2b: State warning confirmation (pause and ask user if any non-Released items found)
    if not state_confirmed:
        state_warnings = collect_state_warnings(tree)
        if state_warnings["obsolete"] or state_warnings["other"]:
            return {
                "status": "needs_state_confirmation",
                "warnings": state_warnings,
            }

    # Step 3: Create Items for all nodes
    item_counters = {"created": 0, "existing": 0, "updated": 0, "failed": 0}
    ensure_items_for_all_nodes(tree, ws, image_loader, item_counters)

    # Step 4: Loose item check
    loose_blocked = _check_loose_items(tree, doc.project, machine_code)
    if loose_blocked:
        items_list = "<br>".join([f"• <b>{ic}</b>" for ic in loose_blocked])
        frappe.throw(
            f"The following items are marked as loose items and cannot have a BOM created.<br>"
            f"Enable 'Can be converted to BOM' in their Project Component Master first:<br><br>"
            f"{items_list}",
            title=_("Loose Items Blocking Upload")
        )

    # Step 5: Create Component Masters
    cm_counters = create_component_masters_for_all_items(tree, doc.project, machine_code)

    # Pre-load CM BOM hashes for this machine (used in Steps 6 and 7).
    # Only items with an existing active_bom are eligible for retain —
    # a CM whose hash was pre-set at creation time (Step 5) but has no active_bom
    # means no BOM exists yet and must not be treated as retained.
    cm_bom_hashes = {}
    if machine_code:
        cm_records = frappe.get_all(
            "Project Component Master",
            filters={"project": doc.project, "machine_code": machine_code, "has_bom": 1},
            fields=["item_code", "bom_structure_hash", "active_bom"],
        )
        cm_bom_hashes = {
            r.item_code: r.bom_structure_hash
            for r in cm_records
            if r.bom_structure_hash and r.active_bom
        }

    # Step 6: Check for BOM version changes (skip if user already confirmed).
    # Items where CM hash == Excel hash are retain candidates — the global BOM may
    # differ (another machine created a newer version) but THIS machine's design
    # hasn't changed, so exclude them from the confirmation dialog.
    if not confirmed:
        version_changes = _scan_for_bom_version_changes(tree, cm_bom_hashes=cm_bom_hashes)
        if version_changes:
            return {
                "status": "needs_confirmation",
                "version_changes": version_changes,
            }

    # Step 7: Create BOMs bottom-up

    bom_counters = _create_boms_for_tree(tree, doc.project, ws, image_loader, cm_bom_hashes=cm_bom_hashes)
    retained_items = bom_counters.pop("retained_items", set())
    reuse_boms = bom_counters.pop("reuse_boms", {})

    # Step 8: Link BOMs to CMs, populate hierarchy, recalculate
    _link_boms_to_component_masters(doc.project, retained_items=retained_items, machine_code=machine_code, reuse_boms=reuse_boms)
    _populate_hierarchy_codes(doc.project, machine_code=machine_code, only_missing=True)
    _refresh_bom_usage_hierarchy_codes(doc.project, machine_code=machine_code)

    from clevertech.project_component_master.bom_hooks import recalculate_component_masters_for_project
    recalc_result = recalculate_component_masters_for_project(doc.project)
    if recalc_result.get("errors"):
        bom_counters["errors"].extend(recalc_result["errors"])

    # Step 9: Log upload history
    doc.append("upload_history", {
        "bom_file": doc.bom_file,
        "machine_code": machine_code,
        "uploaded_by": frappe.session.user,
        "uploaded_on": frappe.utils.now_datetime(),
        "status": "Success",
        "items_created": item_counters.get("created", 0),
        "boms_created": bom_counters.get("created", 0),
        "cms_created": cm_counters.get("created", 0),
    })
    doc.save(ignore_permissions=True)
    frappe.db.commit()

    item_counters["total"] = sum(item_counters.get(k, 0) for k in ("created", "existing", "updated", "failed"))
    cm_counters["total"] = sum(cm_counters.get(k, 0) for k in ("created", "existing", "updated", "failed"))

    return {
        "status": "success",
        "summary": {
            "items": item_counters,
            "boms": {
                "created": bom_counters["created"],
                "existing": bom_counters["skipped"],
                "failed": bom_counters["failed"],
                "total": bom_counters["created"] + bom_counters["skipped"] + bom_counters["failed"],
            },
            "component_masters": cm_counters,
        },
        "errors": bom_counters["errors"],
    }


# ==================== Helpers ====================

def _check_loose_items(tree, project, machine_code):
    """
    Check all nodes (assemblies + leaf items) for existing PCMs marked as
    loose items without can_be_converted_to_bom enabled.
    A loose leaf item blocks its parent assembly's BOM creation.

    Returns:
        list of blocked item codes
    """
    blocked = []

    for node in _get_all_nodes(tree):
        result = frappe.db.get_value(
            "Project Component Master",
            {
                "project": project,
                "item_code": node["item_code"],
                "machine_code": machine_code,
            },
            ["is_loose_item", "can_be_converted_to_bom"],
            as_dict=True,
        )

        if result and result.is_loose_item and not result.can_be_converted_to_bom:
            blocked.append(node["item_code"])

    return blocked


def _scan_for_bom_version_changes(tree, cm_bom_hashes=None):
    """
    Walk tree and find assembly nodes where an active BOM already exists
    but the structure has changed (hash mismatch).

    Args:
        cm_bom_hashes: Optional dict {item_code: bom_structure_hash} for the uploading
            machine's CMs (only items with active_bom set). If the CM hash matches the
            Excel hash, this machine's BOM hasn't changed — skip it even if the global
            BOM differs (another machine may have created a newer version).

    Returns:
        list of dicts: [{"item_code", "description", "existing_bom"}, ...]
    """
    if cm_bom_hashes is None:
        cm_bom_hashes = {}

    changes = []

    def scan_node(node):
        if not node.get("children"):
            return

        new_hash = _calculate_tree_hash(node["children"])

        # If this machine's CM hash matches Excel → retain, not a version change
        if cm_bom_hashes.get(node["item_code"]) == new_hash:
            for child in node.get("children", []):
                scan_node(child)
            return

        # Check if ANY active BOM (default or non-default) already has this structure.
        # If yes → not a version change, no dialog needed.
        # If no → flag as version change only when a default BOM exists (show "changing from X").
        matching_bom = frappe.db.get_value(
            "BOM",
            {"item": node["item_code"], "is_active": 1, "docstatus": 1,
             "custom_bom_structure_hash": new_hash},
            "name",
            order_by="creation desc",
        )

        if not matching_bom:
            existing_default = frappe.db.get_value(
                "BOM",
                {"item": node["item_code"], "is_active": 1, "is_default": 1, "docstatus": 1},
                "name",
            )
            if existing_default:
                changes.append({
                    "item_code": node["item_code"],
                    "description": node.get("description") or "",
                    "existing_bom": existing_default,
                })
        # else: Excel matches an existing active BOM → not a version change, no dialog

        for child in node.get("children", []):
            scan_node(child)

    for root_node in tree:
        scan_node(root_node)

    return changes


def _create_boms_for_tree(tree, project, ws, image_loader, cm_bom_hashes=None):
    """
    Walk tree bottom-up and create BOMs for all assembly nodes.
    create_bom_recursive is idempotent: skips if hash unchanged,
    creates new version if hash differs (ERPNext auto-demotes old).

    Args:
        cm_bom_hashes: Optional dict {item_code: bom_structure_hash} pre-loaded from the
            uploading machine's Component Masters. Passed through to create_bom_recursive
            to enable the machine-level retain check.

    Returns:
        dict: {"created": N, "skipped": N, "failed": N, "errors": [...], "retained_items": set()}
    """
    retained_items = set()
    reuse_boms = {}  # {item_code: bom_name} for items where a non-default active BOM was matched
    counters = {"created": 0, "skipped": 0, "failed": 0, "errors": [], "retained_items": retained_items, "reuse_boms": reuse_boms}

    def process_node(node):
        if not node.get("children"):
            return  # Leaf node — no BOM needed

        # Bottom-up: children before parent
        for child in node["children"]:
            process_node(child)

        try:
            result = create_bom_recursive(node, project, ws, image_loader, cm_bom_hashes=cm_bom_hashes)
            if result == "retain":
                retained_items.add(node["item_code"])
                counters["skipped"] += 1
            elif isinstance(result, tuple) and result[0] == "reuse":
                reuse_boms[node["item_code"]] = result[1]
                counters["skipped"] += 1
            elif result:
                counters["created"] += 1
            else:
                counters["skipped"] += 1
        except Exception as ex:
            counters["failed"] += 1
            counters["errors"].append(f"{node['item_code']}: {str(ex)}")
            frappe.log_error(
                title=f"Phase 1 BOM creation failed: {node['item_code']}",
                message=frappe.get_traceback(),
            )

    for root_node in tree:
        process_node(root_node)

    return counters
