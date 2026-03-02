"""
Enhanced BOM Upload with Validation and Component Master Integration

New "Create BOMs with Validation" button on BOM Upload form.
Imports existing functions from bom_upload.py — does NOT modify that file.

Sequence (Decision 11 & 12):
  1. Parse Excel (reuse parse_rows, build_tree)
  2. Create Items for assemblies (reuse ensure_item_exists)
  3. Create Component Masters FIRST (active_bom=null)
  4. Analyze upload — hash comparison, blocking checks
  5. Create BOMs bottom-up (reuse create_bom_recursive)
  6. Link active_bom back to Component Masters
  7. BOM on_submit hooks auto-populate BOM Usage tables
"""

import frappe
from frappe import _
import openpyxl
import io
import hashlib
import json

from clevertech.clevertech.doctype.bom_upload.bom_upload import (
    parse_rows,  # Original - will be replaced by parse_rows_dynamic for enhanced upload
    build_tree,
    ensure_item_exists,
    create_bom_recursive,
    HAS_IMAGE_LOADER,
    # Utility functions (Decision 12A - reuse from original)
    clean_code,
    to_float,
    normalize_uom,
    # Mapping functions (Decision 12A - auto-inherit when migrated to doctypes)
    normalize_material,
    get_surface_treatment,
    get_type_of_material,
    get_item_group_and_hsn,
)

# Conditional import for image loader (same as bom_upload.py)
try:
    from openpyxl_image_loader import SheetImageLoader
except ImportError:
    SheetImageLoader = None

from clevertech.project_component_master.bom_hooks import (
    calculate_bom_structure_hash as calculate_bom_doc_hash,
)

from clevertech.project_component_master.bulk_generation import (
    populate_bom_usage_tables,
)


# ==================== State → design_status Mapping ====================

# Only three valid STATE values in the Excel (case-insensitive).
# Blank STATE is treated conservatively as "In Creation".
_STATE_TO_DESIGN_STATUS = {
    "released":    "Released",
    "in creation": "In Creation",
    "obsolete":    "Obsolete",
}


def _map_state_to_design_status(state):
    """Return a valid Project Component Master design_status for an Excel STATE value.

    Case-insensitive. Blank or unrecognised value → 'In Creation' (conservative).
    """
    if not state or not str(state).strip():
        return "In Creation"
    return _STATE_TO_DESIGN_STATUS.get(str(state).strip().lower(), "In Creation")


def collect_state_warnings(tree):
    """
    Scan the filtered tree (after G-code filter) for items with notable states.
    Used in Phase 1 to produce non-blocking warnings before upload proceeds.

    Returns:
        dict:
          "obsolete" — items whose STATE is "Obsolete"
          "other"    — items whose STATE is not "Released" and not blank
    """
    warnings = {"obsolete": [], "other": []}

    def _scan(nodes):
        for node in nodes:
            item_code = node.get("item_code", "")
            raw_state = str(node.get("state") or "").strip()
            normalised = raw_state.lower()

            if normalised == "obsolete":
                warnings["obsolete"].append({
                    "item_code": item_code,
                    "description": node.get("description") or "",
                })
            elif normalised and normalised != "released":
                warnings["other"].append({
                    "item_code": item_code,
                    "description": node.get("description") or "",
                    "state": raw_state,
                })

            if node.get("children"):
                _scan(node["children"])

    _scan(tree)
    return warnings


# ==================== Dynamic Column Mapping (Phase 4H) ====================

def map_excel_columns(ws):
    """
    Dynamically map Excel column letters to field names by searching row 2 headers.
    Handles both old and new PE2 export formats where columns may shift.

    Root Cause (2026-02-01): PE2 export format changed, LivelloBom moved from AR to AQ.
    Hardcoded column AR in parse_rows() caused wrong hierarchy (raw materials got BOMs).

    Args:
        ws: Excel worksheet

    Returns:
        dict: Mapping of field names to column letters
              Example: {"item_code": "C", "level": "AQ", ...}

    Raises:
        frappe.ValidationError: If any required column header is not found
    """
    # Define required column headers (exact matches from Excel row 2)
    # Based on RCA 2026-02-01: All headers consistent except LivelloBom position
    header_mapping = {
        # Field name: [List of possible header values in row 2]
        "position": ["Position"],
        "item_code": ["Item no"],
        "description": ["Description"],
        "qty": ["Qty"],
        "revision": ["Rev."],
        "extended_description": ["DESCRIZIONE_ESTESA"],
        "material": ["MATERIAL"],
        "part_number": ["Part_number"],
        "weight": ["WEIGHT"],
        "manufacturer": ["MANUFACTURER"],
        "treatment": ["TIPO_TRATTAMENTO"],
        "uom": ["UM"],
        "level": ["LivelloBom"],  # CRITICAL: Shifted between formats (AR vs AQ)
    }

    # Search row 2 for headers
    found_columns = {}
    missing_columns = []

    # Read all headers from row 2 (up to column 100)
    headers_in_excel = {}
    for col_num in range(1, 100):
        col_letter = openpyxl.utils.get_column_letter(col_num)
        header_val = ws[f"{col_letter}2"].value
        if header_val:
            headers_in_excel[col_letter] = str(header_val).strip()

    # Map each required field to its column
    for field_name, possible_headers in header_mapping.items():
        found = False
        for col_letter, header_val in headers_in_excel.items():
            if header_val in possible_headers:
                found_columns[field_name] = col_letter
                found = True
                break

        if not found:
            missing_columns.append({
                "field": field_name,
                "expected_headers": possible_headers
            })

    # Validate all required columns are present
    if missing_columns:
        error_msg = _("Excel file format validation failed. Missing required columns:")
        error_msg += "\n\n"
        for missing in missing_columns:
            error_msg += f"• Field: {missing['field']}\n"
            error_msg += f"  Expected header: {', '.join(missing['expected_headers'])}\n\n"

        error_msg += f"\nFound headers in row 2:\n"
        for col, header in sorted(headers_in_excel.items())[:30]:  # Show first 30 columns
            error_msg += f"  Column {col}: {header}\n"

        frappe.throw(error_msg, title=_("Invalid Excel Format"))

    return found_columns


def parse_rows_dynamic(ws):
    """
    Parse Excel rows using dynamic column mapping (handles format changes).
    Replacement for hardcoded parse_rows() from bom_upload.py (Phase 4H).

    Uses map_excel_columns() to find column positions by header name instead of
    hardcoded positions. This handles PE2 export format changes automatically.

    Args:
        ws: Excel worksheet

    Returns:
        list: List of row dicts with parsed data (same format as parse_rows)
    """
    # Get dynamic column mapping (validates required columns exist)
    col_map = map_excel_columns(ws)

    rows = []

    # Start from row 3 (row 1 = title, row 2 = headers, row 3+ = data)
    for r in range(3, ws.max_row + 1):
        # Check if item_code exists (required field)
        code_raw = ws[f"{col_map['item_code']}{r}"].value
        if not code_raw:
            continue  # Skip empty rows

        # Read position (column A or B fallback, same as original)
        position_col = col_map.get("position", "A")
        position = ws[f"{position_col}{r}"].value or ws[f"B{r}"].value

        rows.append({
            "row_num": r,  # Store row number for image extraction
            "position": position,
            "item_code": clean_code(code_raw),
            "description": ws[f"{col_map['description']}{r}"].value,
            "extended_description": ws[f"{col_map['extended_description']}{r}"].value,
            "qty": to_float(ws[f"{col_map['qty']}{r}"].value, 1),
            "revision": ws[f"{col_map['revision']}{r}"].value,
            "material": ws[f"{col_map['material']}{r}"].value,
            "part_number": ws[f"{col_map['part_number']}{r}"].value,
            "weight": to_float(ws[f"{col_map['weight']}{r}"].value, 0),
            "manufacturer": ws[f"{col_map['manufacturer']}{r}"].value,
            "treatment": ws[f"{col_map['treatment']}{r}"].value,
            "uom": normalize_uom(ws[f"{col_map['uom']}{r}"].value),
            "level": int(ws[f"{col_map['level']}{r}"].value or 0),  # CRITICAL!
            "state": ws[f"J{r}"].value,  # Design state from column J (G-code release status)
            "children": []
        })

    return rows


# ==================== Main Entry Point ====================

@frappe.whitelist()
def create_boms_with_validation(docname):
    """
    Enhanced BOM creation with validation and Component Master integration.

    Args:
        docname: BOM Upload document name

    Returns:
        dict: Analysis results or creation summary
    """
    doc = frappe.get_doc("BOM Upload", docname)

    if not doc.bom_file:
        frappe.throw(_("Please attach a BOM Excel file first."))

    if not doc.project:
        frappe.throw(_("Please select a Project first."))

    # Validate machine_code (required for new uploads)
    machine_code = doc.machine_code
    if not machine_code:
        frappe.throw(_("Machine Code is required for BOM Upload. Please enter the machine code (e.g., P0000000003033)."))

    # Step 1: Parse Excel (reuse existing logic)
    file_doc = frappe.get_doc("File", {"file_url": doc.bom_file})
    content = file_doc.get_content()

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Initialize image loader only if available
    image_loader = None
    if HAS_IMAGE_LOADER and SheetImageLoader:
        try:
            image_loader = SheetImageLoader(ws)
        except Exception:
            pass  # Image loading optional

    rows = parse_rows_dynamic(ws)  # Phase 4H: Dynamic column mapping
    tree = build_tree(rows)

    if not tree:
        frappe.throw(_("No components found in the Excel file."))

    # Filter tree to remove non-released G-codes (STATE != "RELEASED")
    filtered_tree, skipped_info = filter_tree_by_g_code_state(tree)

    # Notify user if any G-codes were skipped
    if skipped_info["count"] > 0:
        skip_details = "<br>".join([
            f"• <b>{item['item_code']}</b> ({item['description'][:50]}...) - STATE: {item['state']} - Skipped {item['subtree_count']} items"
            for item in skipped_info["items"]
        ])
        frappe.msgprint(
            _(
                f"<b>Skipped {skipped_info['count']} non-released G-code(s)</b><br><br>"
                f"Total items skipped (including children): <b>{skipped_info['total_items_skipped']}</b><br><br>"
                f"{skip_details}<br><br>"
                "<small><i class='fa fa-info-circle'></i> Only G-codes with STATE = 'RELEASED' are processed. "
                "Non-released designs and their child components are excluded from upload.</small>"
            ),
            title=_("Non-Released G-Codes Skipped"),
            indicator="orange"
        )

    # Use filtered tree for all subsequent operations
    tree = filtered_tree

    if not tree:
        frappe.throw(_("No releasable components found in the Excel file after STATE filtering."))

    # Step 2: Create Items for ALL nodes (assemblies + leaf items)
    # Items must exist BEFORE Component Masters (item_code is a Link field to Item)
    item_counters = {"created": 0, "existing": 0, "updated": 0, "failed": 0}
    ensure_items_for_all_nodes(tree, ws, image_loader, item_counters)

    # Step 3: Create Component Masters FIRST for ALL items (assemblies + raw materials)
    # All items default to "Buy" (conservative). T, Y, E items get released_for_procurement=No.
    # (active_bom=null at this stage for assemblies, set later after BOM creation)
    cm_counters = create_component_masters_for_all_items(tree, doc.project, machine_code)

    # Step 4: Analyze upload (now Component Masters exist for blocking checks)
    analysis = analyze_upload(tree, doc.project, machine_code)

    # Step 5: Check for blocking issues
    if analysis["loose_blocked"]:
        return {
            "status": "blocked",
            "reason": "loose_items_not_enabled",
            "analysis": _serialize_analysis(analysis),
            "message": _("Enable 'Can be converted to BOM' for loose items first"),
        }

    if analysis["changed_components"]:
        # Categorize changed components by blocking level
        hard_blocked = []  # MR/RFQ blocks - user must deactivate old BOM
        manager_blocked = []  # PO blocks - need manager role
        confirmable = []  # No procurement - just need confirmation

        for comp in analysis["changed_components"]:
            details = comp.get("details", {})
            blocking_level = details.get("blocking_level", "confirm")

            if blocking_level == "manager_required":
                if details.get("can_proceed"):
                    # User has manager role, can confirm
                    confirmable.append(comp)
                else:
                    # User lacks manager role
                    manager_blocked.append(comp)
            elif blocking_level == "block":
                hard_blocked.append(comp)
            else:
                # confirm level
                confirmable.append(comp)

        if hard_blocked:
            return {
                "status": "procurement_blocked",
                "reason": "active_mr_rfq",
                "blocked_components": [
                    {
                        "item_code": c["node"]["item_code"],
                        "details": c.get("details", {}),
                    }
                    for c in hard_blocked
                ],
                "analysis": _serialize_analysis(analysis),
                "message": _("Cannot change BOM. Deactivate old BOMs manually, then re-run upload."),
            }

        if manager_blocked:
            return {
                "status": "manager_required",
                "reason": "active_po_no_role",
                "blocked_components": [
                    {
                        "item_code": c["node"]["item_code"],
                        "details": c.get("details", {}),
                    }
                    for c in manager_blocked
                ],
                "analysis": _serialize_analysis(analysis),
                "message": _("Cannot change BOM. Child items have active POs. Manager role required."),
            }

        if confirmable:
            return {
                "status": "requires_confirmation",
                "confirmable_components": [
                    {
                        "item_code": c["node"]["item_code"],
                        "details": c.get("details", {}),
                    }
                    for c in confirmable
                ],
                "analysis": _serialize_analysis(analysis),
                "message": _("BOM structure changed. Confirm with remarks to proceed."),
            }

    # Step 6 & 7: Create BOMs and link back to Component Masters
    result = create_boms_and_link_components(tree, doc.project, analysis, ws, image_loader, machine_code)

    # Build comprehensive summary
    result["summary"] = {
        "items": {
            "created": item_counters["created"],
            "existing": item_counters["existing"],
            "updated": item_counters["updated"],
            "failed": item_counters["failed"],
            "total": sum(item_counters.values())
        },
        "boms": {
            "created": result.pop("created", 0),
            "existing": result.pop("skipped", 0),
            "failed": result.pop("failed", 0),
            "total": 0  # Will be calculated below
        },
        "component_masters": {
            "created": cm_counters["created"],
            "existing": cm_counters["existing"],
            "updated": cm_counters["updated"],
            "failed": cm_counters["failed"],
            "total": sum(cm_counters.values())
        }
    }
    result["summary"]["boms"]["total"] = (
        result["summary"]["boms"]["created"] +
        result["summary"]["boms"]["existing"] +
        result["summary"]["boms"]["failed"]
    )

    # Ensure status is set explicitly
    result["status"] = "success"

    # Build user-friendly message
    result["message"] = _build_summary_message(result["summary"], result.get("errors", []))

    # Log upload history
    doc.append("upload_history", {
        "bom_file": doc.bom_file,
        "machine_code": machine_code,
        "uploaded_by": frappe.session.user,
        "uploaded_on": frappe.utils.now_datetime(),
        "status": "Success",
        "items_created": result["summary"]["items"].get("created", 0),
        "boms_created": result["summary"]["boms"].get("created", 0),
        "cms_created": result["summary"]["component_masters"].get("created", 0),
    })
    doc.save(ignore_permissions=True)

    frappe.log_error(title="BOM Upload Debug - Fresh Upload", message=str(result))
    frappe.db.commit()
    return result


# ==================== Step 2: Ensure Items ====================

def ensure_items_for_all_nodes(tree, ws, image_loader, counters=None):
    """
    Walk tree and create Item masters for ALL nodes (assemblies + leaf items).
    Must run BEFORE Component Master creation since item_code is a Link field.
    Uses ensure_item_exists() from existing BOM Upload module (idempotent).

    Args:
        tree: List of root nodes
        ws: Excel worksheet
        image_loader: SheetImageLoader instance (or None)
        counters: dict to track {"created": 0, "existing": 0, "failed": 0}

    Returns:
        dict: counters
    """
    if counters is None:
        counters = {"created": 0, "existing": 0, "failed": 0}

    for node in tree:
        result = ensure_item_exists(
            node["item_code"],
            node.get("description"),
            node.get("extended_description"),
            node.get("uom"),
            node.get("row_num"),
            ws,
            image_loader,
            node.get("material"),
            node.get("treatment"),
            node.get("weight"),
            node.get("part_number"),
            node.get("manufacturer"),
            node.get("revision")
        )
        counters[result] = counters.get(result, 0) + 1

        if node.get("children"):
            ensure_items_for_all_nodes(node["children"], ws, image_loader, counters)

    return counters


# ==================== Step 3: Create Component Masters ====================

def create_component_masters_for_all_items(tree, project, machine_code):
    """
    Create Component Masters for ALL nodes (assemblies + leaf/RM items).
    For existing CMs: apply make/buy merge logic (Excel wins if present, else keep existing).
    For new CMs: set defaults based on item type.

    Business Rules:
    - All items default to "Buy" (conservative approach)
    - T, Y, E prefixed items: released_for_procurement = "No"
    - Other items: released_for_procurement = "Yes"
    - machine_code is set on all Component Masters (enables per-machine tracking)

    Args:
        tree: List of root nodes
        project: Project name
        machine_code: Machine code from BOM Upload (e.g., P0000000003033)

    Returns:
        dict: {"created": X, "existing": X, "updated": X, "failed": X}
    """
    counters = {"created": 0, "existing": 0, "updated": 0, "failed": 0}
    all_nodes = _get_all_nodes(tree)

    # Look up Cost Center linked to this machine_code (one query for all CMs)
    cost_center = frappe.db.get_value("Cost Center", {"custom_machine_code": machine_code}, "name")

    for node in all_nodes:
        item_code = node["item_code"]
        is_assembly = bool(node.get("children"))

        # Check if Component Master already exists (project + item_code + machine_code unique)
        existing = frappe.db.exists(
            "Project Component Master",
            {
                "project": project,
                "item_code": item_code,
                "machine_code": machine_code
            },
        )

        if existing:
            # Merge logic: update make_or_buy only if Excel has a value
            updated = _merge_make_or_buy(existing, node)
            if updated:
                counters["updated"] += 1
            else:
                counters["existing"] += 1
            continue

        try:
            # Check if item starts with T, Y, or E
            item_code_upper = str(item_code).upper()
            if item_code_upper.startswith(("T", "Y", "E")):
                released_for_procurement = "No"
            else:
                released_for_procurement = "Yes"

            # Only set project_qty for root assemblies (level 1)
            # Child items (level 2+) get project_qty=0, their requirement comes from bom_qty_required
            level = node.get("level", 0)
            project_qty = node.get("qty", 0) if level == 1 else 0

            # Map STATE to design_status (case-insensitive; blank → "In Creation")
            design_status = _map_state_to_design_status(node.get("state"))

            cm_data = {
                "doctype": "Project Component Master",
                "project": project,
                "item_code": item_code,
                "machine_code": machine_code,  # Set machine code on all Component Masters
                "bom_level": level,
                "is_loose_item": 0,
                "design_status": design_status,  # Mapped from Excel STATE column
                "project_qty": project_qty,  # From Excel column E (root assembly only)
                "created_from": "BOM Upload",
                "released_for_procurement": released_for_procurement,
                "cost_center": cost_center,  # Linked via machine_code on Cost Center
            }

            # Make/Buy: use Excel value if present, otherwise leave blank for user to set manually
            if is_assembly:
                cm_data["has_bom"] = 1
                cm_data["active_bom"] = None  # Set later after BOM creation
                cm_data["bom_structure_hash"] = calculate_tree_structure_hash(node["children"])
                cm_data["make_or_buy"] = node.get("make_or_buy") or ""
            else:
                cm_data["has_bom"] = 0
                cm_data["make_or_buy"] = node.get("make_or_buy") or ""

            cm = frappe.get_doc(cm_data)
            cm.flags.ignore_validate = True
            cm.flags.ignore_mandatory = True
            cm.insert(ignore_permissions=True)
            counters["created"] += 1
        except Exception:
            counters["failed"] += 1
            frappe.log_error(
                title=f"Failed to create Component Master for {item_code}",
                message=frappe.get_traceback(),
            )

    return counters


def _merge_make_or_buy(cm_name, node):
    """
    Merge make/buy from Excel into existing Component Master.
    Only updates if Excel has an explicit value; otherwise keeps existing.

    This allows PE2 team to only tag new items or changes in the Excel.
    Previously tagged items can be left blank.

    Returns:
        bool: True if value was updated, False if unchanged
    """
    excel_make_or_buy = node.get("make_or_buy")
    if not excel_make_or_buy:
        return False  # No value in Excel, keep existing

    current_value = frappe.db.get_value(
        "Project Component Master", cm_name, "make_or_buy"
    )

    if current_value != excel_make_or_buy:
        frappe.db.set_value(
            "Project Component Master", cm_name,
            "make_or_buy", excel_make_or_buy
        )
        return True
    return False


def _get_all_nodes(tree):
    """Flatten tree to get ALL nodes (assemblies + leaf items)."""
    all_nodes = []
    for node in tree:
        all_nodes.append(node)
        if node.get("children"):
            all_nodes.extend(_get_all_nodes(node["children"]))
    return all_nodes


def _get_assembly_nodes(tree):
    """Flatten tree to get all assembly nodes (nodes that have children)."""
    assemblies = []
    for node in tree:
        if node.get("children"):
            assemblies.append(node)
            assemblies.extend(_get_assembly_nodes(node["children"]))
    return assemblies


def filter_tree_by_g_code_state(tree):
    """
    Filter tree to remove G-code items with non-released STATE.

    Business Rule (Approach 1 - G-level only):
    - If item starts with "G" and STATE != "RELEASED" (or blank), skip entire subtree
    - This prevents creation of Items, Component Masters, and BOMs for unreleased designs

    Args:
        tree: List of root nodes from build_tree()

    Returns:
        tuple: (filtered_tree, skipped_info)
            filtered_tree: Tree with non-released G-codes removed
            skipped_info: Dict with {"count": N, "items": [...]} for user notification
    """
    filtered_tree = []
    skipped_items = []

    def should_skip_node(node):
        """Check if this G-code node should be skipped."""
        item_code = str(node.get("item_code", "")).upper()

        # Only check G-codes
        if not item_code.startswith("G"):
            return False

        # Get STATE value
        state = node.get("state")

        # Skip if STATE is blank or not "RELEASED"
        if not state or str(state).strip().upper() != "RELEASED":
            return True

        return False

    def count_subtree_nodes(node):
        """Count total nodes in a subtree."""
        count = 1  # Count this node
        if node.get("children"):
            for child in node["children"]:
                count += count_subtree_nodes(child)
        return count

    def filter_recursive(nodes):
        """Recursively filter nodes."""
        filtered = []
        for node in nodes:
            if should_skip_node(node):
                # Skip this G-code and all its children
                item_count = count_subtree_nodes(node)
                skipped_items.append({
                    "item_code": node.get("item_code"),
                    "description": node.get("description"),
                    "state": node.get("state") or "(blank)",
                    "subtree_count": item_count
                })
            else:
                # Keep this node, but filter its children recursively
                filtered_node = node.copy()
                if node.get("children"):
                    filtered_node["children"] = filter_recursive(node["children"])
                filtered.append(filtered_node)
        return filtered

    filtered_tree = filter_recursive(tree)

    skipped_info = {
        "count": len(skipped_items),
        "total_items_skipped": sum(item["subtree_count"] for item in skipped_items),
        "items": skipped_items
    }

    return filtered_tree, skipped_info


# ==================== Step 4: Analyze Upload ====================

def analyze_upload(tree, project, machine_code):
    """
    Analyze upload tree for new, unchanged, changed, and blocked components.

    Args:
        tree: List of root nodes
        project: Project name
        machine_code: Machine code from BOM Upload

    Returns:
        dict with can_create, changed_components, loose_blocked,
        blocked_by_dependencies, summary
    """
    all_assemblies = _get_assembly_nodes(tree)
    graph = _build_dependency_graph(all_assemblies)

    categorized = {
        "new": [],
        "unchanged": [],
        "changed": [],
        "loose_blocked": [],
    }

    for item_code, data in graph.items():
        status, details = _determine_component_status(
            item_code, data["node"], project, machine_code
        )
        data["status"] = status
        data["details"] = details
        categorized[status].append(item_code)

    # Find blocking dependencies (ancestor chain)
    changed_set = set(categorized["changed"])
    loose_blocked_set = set(categorized["loose_blocked"])
    blocked_by_changes = _find_blocked_ancestors(
        graph, changed_set, loose_blocked_set
    )

    # Build list of items safe to create
    blocked_set = set(blocked_by_changes.keys())
    can_create = []
    for item_code in categorized["new"] + categorized["unchanged"]:
        if item_code not in blocked_set:
            can_create.append(graph[item_code]["node"])

    return {
        "can_create": can_create,
        "changed_components": [graph[ic] for ic in categorized["changed"]],
        "loose_blocked": [graph[ic] for ic in categorized["loose_blocked"]],
        "blocked_by_dependencies": blocked_by_changes,
        "summary": {
            "total": len(all_assemblies),
            "new": len(categorized["new"]),
            "unchanged": len(categorized["unchanged"]),
            "changed": len(categorized["changed"]),
            "loose_blocked": len(categorized["loose_blocked"]),
            "can_create": len(can_create),
            "blocked": len(blocked_by_changes),
        },
    }


def _determine_component_status(item_code, node, project, machine_code):
    """
    Determine if component is new, unchanged, changed, or loose_blocked.

    Args:
        item_code: Item code to check
        node: Tree node dict
        project: Project name
        machine_code: Machine code from BOM Upload

    Returns:
        tuple: (status, details)
    """
    component_master = frappe.db.get_value(
        "Project Component Master",
        {
            "project": project,
            "item_code": item_code,
            "machine_code": machine_code
        },
        ["name", "is_loose_item", "can_be_converted_to_bom",
         "bom_structure_hash", "active_bom"],
        as_dict=True,
    )

    if not component_master:
        return "new", {"first_time": True}

    # Check loose item blocking
    if component_master.is_loose_item and not component_master.can_be_converted_to_bom:
        return "loose_blocked", {
            "component_master": component_master.name,
            "message": f"Loose item {item_code} not enabled for BOM conversion",
        }

    # Check for BOM changes — no active BOM linked to Component Master yet
    if not component_master.active_bom:
        # But check if an active BOM already exists in system (might not be linked yet)
        existing_bom = frappe.db.get_value(
            "BOM",
            {
                "item": item_code,
                "project": project,
                "is_active": 1,
                "is_default": 1,
                "docstatus": 1,
            },
            "name",
        )

        if not existing_bom:
            # Also check without project filter (BOM may not have project set)
            existing_bom = frappe.db.get_value(
                "BOM",
                {
                    "item": item_code,
                    "is_active": 1,
                    "is_default": 1,
                    "docstatus": 1,
                },
                "name",
            )

        if not existing_bom:
            # No existing BOM found — truly new
            return "new", {
                "component_master": component_master.name,
                "first_time": False,
            }

        # BOM exists but not linked to Component Master — compare hashes
        children = node.get("children", [])
        new_hash = calculate_tree_structure_hash(children)

        existing_hash = frappe.db.get_value(
            "BOM", existing_bom, "custom_bom_structure_hash"
        )

        # Backfill hash if NULL
        if not existing_hash:
            bom_doc = frappe.get_doc("BOM", existing_bom)
            existing_hash = calculate_bom_doc_hash(bom_doc)
            if existing_hash:
                frappe.db.set_value(
                    "BOM", existing_bom,
                    "custom_bom_structure_hash", existing_hash,
                    update_modified=False
                )

        if new_hash == existing_hash:
            # BOM exists and hash matches — unchanged, will be linked later
            return "unchanged", {
                "component_master": component_master.name,
                "bom": existing_bom,
                "existing_bom_not_linked": True,
            }
        else:
            # BOM exists but hash differs — version change needed
            blocking_info = _check_procurement_blocking(project, item_code, existing_bom)
            children = node.get("children", [])
            bom_diff = calculate_bom_diff(existing_bom, children)
            return "changed", {
                "change_type": "bom_structure",
                "component_master": component_master.name,
                "old_bom": existing_bom,
                "old_hash": existing_hash,
                "new_hash": new_hash,
                "blocking_level": blocking_info["blocking_level"],
                "blocking_message": blocking_info["message"],
                "can_proceed": blocking_info["can_proceed"],
                "procurement_docs": blocking_info["procurement_docs"],
                "existing_bom_not_linked": True,
                "bom_diff": bom_diff,
                "impacted_parents": _get_impacted_parent_boms(existing_bom),
            }

    # Compare BOM structures via hash - get hash from BOM itself (source of truth)
    children = node.get("children", [])
    new_hash = calculate_tree_structure_hash(children)

    # Get existing hash from BOM (not Component Master)
    existing_hash = frappe.db.get_value(
        "BOM", component_master.active_bom, "custom_bom_structure_hash"
    )

    # Backfill if NULL - calculate from actual BOM items
    if not existing_hash:
        bom_doc = frappe.get_doc("BOM", component_master.active_bom)
        existing_hash = calculate_bom_doc_hash(bom_doc)
        if existing_hash:
            # Store the backfilled hash on BOM
            frappe.db.set_value(
                "BOM", component_master.active_bom,
                "custom_bom_structure_hash", existing_hash,
                update_modified=False
            )

    if new_hash != existing_hash:
        # Check procurement blocking level
        blocking_info = _check_procurement_blocking(
            project, item_code, component_master.active_bom
        )
        bom_diff = calculate_bom_diff(component_master.active_bom, children)
        return "changed", {
            "change_type": "bom_structure",
            "component_master": component_master.name,
            "old_bom": component_master.active_bom,
            "old_hash": existing_hash,
            "new_hash": new_hash,
            "blocking_level": blocking_info["blocking_level"],
            "blocking_message": blocking_info["message"],
            "can_proceed": blocking_info["can_proceed"],
            "procurement_docs": blocking_info["procurement_docs"],
            "bom_diff": bom_diff,
            "impacted_parents": _get_impacted_parent_boms(component_master.active_bom),
        }

    return "unchanged", {
        "component_master": component_master.name,
        "bom": component_master.active_bom,
    }


def _build_dependency_graph(assemblies):
    """
    Build a dependency graph from assembly nodes.
    Maps item_code -> {node, children (item_codes), status, details}
    """
    graph = {}
    for node in assemblies:
        child_codes = [c["item_code"] for c in node.get("children", [])]
        graph[node["item_code"]] = {
            "node": node,
            "children": child_codes,
            "status": None,
            "details": None,
        }
    return graph


def _find_blocked_ancestors(graph, changed_set, loose_blocked_set):
    """
    Find all components blocked by changed or loose-blocked children.
    Uses iterative approach to propagate blocking up the tree.

    Returns:
        dict: Mapping of blocked item codes to list of blocking reasons
    """
    blocked = {}
    # Multiple passes to propagate transitive blocking
    changed = True
    while changed:
        changed = False
        for item_code, data in graph.items():
            if item_code in blocked or item_code in changed_set or item_code in loose_blocked_set:
                continue

            blocking_children = []
            for child_code in data["children"]:
                if child_code in changed_set:
                    blocking_children.append({
                        "item": child_code,
                        "reason": "BOM structure changed",
                    })
                elif child_code in loose_blocked_set:
                    blocking_children.append({
                        "item": child_code,
                        "reason": "Loose item conversion not enabled",
                    })
                elif child_code in blocked:
                    blocking_children.append({
                        "item": child_code,
                        "reason": "Depends on blocked component",
                    })

            if blocking_children:
                blocked[item_code] = blocking_children
                changed = True

    return blocked


def _get_procurement_documents(project, item_code):
    """
    Get active procurement documents for a component.
    Returns list of dicts with doctype, name, status.
    """
    cm_name = frappe.db.get_value(
        "Project Component Master",
        {"project": project, "item_code": item_code},
        "name",
    )
    if not cm_name:
        return []

    records = frappe.get_all(
        "Component Procurement Record",
        filters={"parent": cm_name},
        fields=["document_type", "document_name", "status", "quantity"],
    )

    return [
        {
            "doctype": r.document_type,
            "name": r.document_name,
            "status": r.status or "Unknown",
            "quantity": r.quantity,
        }
        for r in records
    ]


# ==================== Impacted Parent BOMs (Where-Used) ====================

def _get_impacted_parent_boms(old_bom_name):
    """
    Find all active submitted parent BOMs whose BOM Item rows reference
    old_bom_name via bom_no. These parents still point to the old (demoted)
    child BOM and will need engineer review for controlled upward propagation.

    Args:
        old_bom_name: Name of the demoted child BOM

    Returns:
        list: [{"parent_bom": str, "parent_item": str}, ...]
    """
    if not old_bom_name:
        return []

    rows = frappe.db.sql("""
        SELECT DISTINCT b.name AS parent_bom, b.item AS parent_item
        FROM `tabBOM Item` bi
        INNER JOIN `tabBOM` b ON b.name = bi.parent
        WHERE bi.bom_no = %(old_bom)s
          AND b.docstatus = 1
          AND b.is_active = 1
    """, {"old_bom": old_bom_name}, as_dict=True)

    return [{"parent_bom": r.parent_bom, "parent_item": r.parent_item} for r in rows]


# ==================== Phase 4E: Procurement Blocking ====================

def _get_material_requests(project, item_code):
    """Get active (non-cancelled) Material Requests for the item in the project."""
    return frappe.db.sql("""
        SELECT
            mr.name, mr.docstatus, mr.status,
            mri.qty
        FROM `tabMaterial Request` mr
        INNER JOIN `tabMaterial Request Item` mri ON mri.parent = mr.name
        WHERE mri.item_code = %(item_code)s
        AND mr.custom_project_ = %(project)s
        AND mr.docstatus != 2
    """, {"item_code": item_code, "project": project}, as_dict=True)


def _get_rfqs(project, item_code):
    """Get active (non-cancelled) RFQs for the item in the project."""
    return frappe.db.sql("""
        SELECT
            rfq.name, rfq.docstatus, rfq.status,
            rfqi.qty
        FROM `tabRequest for Quotation` rfq
        INNER JOIN `tabRequest for Quotation Item` rfqi ON rfqi.parent = rfq.name
        WHERE rfqi.item_code = %(item_code)s
        AND rfq.custom_project = %(project)s
        AND rfq.docstatus != 2
    """, {"item_code": item_code, "project": project}, as_dict=True)


def _get_purchase_orders(project, item_code):
    """Get active (non-cancelled) Purchase Orders for the item in the project."""
    return frappe.db.sql("""
        SELECT
            po.name, po.docstatus, po.status,
            poi.qty
        FROM `tabPurchase Order` po
        INNER JOIN `tabPurchase Order Item` poi ON poi.parent = po.name
        WHERE poi.item_code = %(item_code)s
        AND po.project = %(project)s
        AND po.docstatus != 2
    """, {"item_code": item_code, "project": project}, as_dict=True)


def _max_blocking_level(current, new_level):
    """Return the stricter blocking level. Order: none < confirm < block < manager_required."""
    levels = {"none": 0, "confirm": 1, "block": 2, "manager_required": 3}
    if levels.get(new_level, 0) > levels.get(current, 0):
        return new_level
    return current


def _can_override_po_block():
    """Check if current user has manager role for PO-stage override."""
    user_roles = frappe.get_roles()
    return "Component Master Manager" in user_roles or "System Manager" in user_roles


def _check_procurement_blocking(project, item_code, old_bom_name):
    """
    Check if procurement exists and determine blocking level.

    Checks children of the old BOM for active MR/RFQ/PO.

    Returns:
        dict: {
            "can_proceed": bool,
            "blocking_level": "none" | "confirm" | "block" | "manager_required",
            "procurement_docs": [...],
            "message": str
        }
    """
    if not old_bom_name:
        return {
            "can_proceed": True,
            "blocking_level": "none",
            "procurement_docs": [],
            "message": "No existing BOM",
        }

    # Get all child items from old BOM
    old_bom_items = frappe.get_all(
        "BOM Item",
        filters={"parent": old_bom_name},
        fields=["item_code"]
    )

    blocking_level = "none"
    procurement_docs = []

    # Check each child item for procurement
    for bom_item in old_bom_items:
        child_item = bom_item.item_code

        # Check Material Requests
        mrs = _get_material_requests(project, child_item)
        if mrs:
            for mr in mrs:
                procurement_docs.append({
                    "doctype": "Material Request",
                    "name": mr.name,
                    "status": mr.status or "Draft" if mr.docstatus == 0 else "Submitted",
                    "quantity": mr.qty,
                    "item_code": child_item,
                })
            # MR exists - confirmable (changed from "block" to allow confirmation)
            blocking_level = _max_blocking_level(blocking_level, "confirm")

        # Check RFQs
        rfqs = _get_rfqs(project, child_item)
        if rfqs:
            for rfq in rfqs:
                procurement_docs.append({
                    "doctype": "Request for Quotation",
                    "name": rfq.name,
                    "status": rfq.status or "Draft" if rfq.docstatus == 0 else "Submitted",
                    "quantity": rfq.qty,
                    "item_code": child_item,
                })
            # RFQ exists - confirmable (changed from "block" to allow confirmation)
            blocking_level = _max_blocking_level(blocking_level, "confirm")

        # Check Purchase Orders
        pos = _get_purchase_orders(project, child_item)
        if pos:
            for po in pos:
                procurement_docs.append({
                    "doctype": "Purchase Order",
                    "name": po.name,
                    "status": po.status or "Draft" if po.docstatus == 0 else "Submitted",
                    "quantity": po.qty,
                    "item_code": child_item,
                })
            blocking_level = _max_blocking_level(blocking_level, "manager_required")

    # If no procurement found, still need confirmation for hash change
    if blocking_level == "none":
        blocking_level = "confirm"

    # Determine if user can proceed
    can_proceed = False
    message = ""

    if blocking_level == "confirm":
        can_proceed = True
        message = "BOM structure changed. Confirmation with remarks required."
    elif blocking_level == "block":
        can_proceed = False
        message = "Cannot change BOM. Child items have active Material Requests or RFQs."
    elif blocking_level == "manager_required":
        if _can_override_po_block():
            can_proceed = True
            message = "Child items have active Purchase Orders. Manager override available."
        else:
            can_proceed = False
            message = "Cannot change BOM. Child items have active Purchase Orders. Manager role required."

    return {
        "can_proceed": can_proceed,
        "blocking_level": blocking_level,
        "procurement_docs": procurement_docs,
        "message": message,
    }


@frappe.whitelist()
def confirm_version_change(docname, confirmations):
    """
    Confirm BOM version changes and proceed with upload.

    Args:
        docname: BOM Upload document name
        confirmations: JSON string of list of dicts with item_code and remarks

    Returns:
        dict: Creation summary
    """
    if isinstance(confirmations, str):
        confirmations = json.loads(confirmations)

    doc = frappe.get_doc("BOM Upload", docname)
    project = doc.project
    machine_code = doc.machine_code

    if not machine_code:
        frappe.throw(_("Machine Code is required for BOM Upload"))

    # Store remarks in Component Masters
    for conf in confirmations:
        item_code = conf.get("item_code")
        remarks = conf.get("remarks", "")

        cm_name = frappe.db.get_value(
            "Project Component Master",
            {
                "project": project,
                "item_code": item_code,
                "machine_code": machine_code
            },
            "name"
        )

        if cm_name:
            frappe.db.set_value(
                "Project Component Master",
                cm_name,
                "version_change_remarks",
                remarks
            )

    # Re-run the upload, now allowing changed components
    return _proceed_with_confirmed_changes(doc, [c["item_code"] for c in confirmations])


def _proceed_with_confirmed_changes(doc, confirmed_items):
    """
    Proceed with BOM creation for confirmed changed items.
    Deactivates old BOMs and creates new ones.
    """
    # Get machine_code from doc
    machine_code = doc.machine_code
    if not machine_code:
        frappe.throw(_("Machine Code is required for BOM Upload"))

    # Parse Excel again
    file_doc = frappe.get_doc("File", {"file_url": doc.bom_file})
    content = file_doc.get_content()

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Initialize image loader
    image_loader = None
    if HAS_IMAGE_LOADER and SheetImageLoader:
        try:
            image_loader = SheetImageLoader(ws)
        except Exception:
            pass

    rows = parse_rows_dynamic(ws)  # Phase 4H: Dynamic column mapping
    tree = build_tree(rows)

    if not tree:
        frappe.throw(_("No components found in the Excel file."))

    project = doc.project

    # Ensure items exist
    item_counters = {"created": 0, "existing": 0, "updated": 0, "failed": 0}
    ensure_items_for_all_nodes(tree, ws, image_loader, item_counters)

    # Track old BOM names for confirmed items so we can report impacted parent BOMs.
    # NOTE: We do NOT manually demote old BOMs here. When new BOM is submitted with
    # is_default=1, ERPNext automatically demotes the old BOM (sets is_default=0).
    # See erpnext/manufacturing/doctype/bom/bom.py - only one BOM can be default per item.
    old_boms = {}  # item_code -> old_bom_name
    for item_code in confirmed_items:
        cm = frappe.db.get_value(
            "Project Component Master",
            {
                "project": project,
                "item_code": item_code,
                "machine_code": machine_code
            },
            ["name", "active_bom"],
            as_dict=True
        )
        if cm and cm.active_bom:
            old_boms[item_code] = cm.active_bom
        elif cm:
            # CM exists but active_bom not linked yet — look up default BOM from system
            existing_bom = frappe.db.get_value(
                "BOM",
                {"item": item_code, "is_active": 1, "is_default": 1, "docstatus": 1},
                "name"
            )
            if existing_bom:
                old_boms[item_code] = existing_bom

    # Build analysis with confirmed items treated as "new"
    all_assemblies = _get_assembly_nodes(tree)
    confirmed_set = set(confirmed_items)

    can_create = []
    for node in all_assemblies:
        item_code = node["item_code"]

        # Check if already has active BOM (unchanged)
        cm = frappe.db.get_value(
            "Project Component Master",
            {
                "project": project,
                "item_code": item_code,
                "machine_code": machine_code
            },
            ["active_bom", "bom_structure_hash"],
            as_dict=True
        )

        if cm and cm.active_bom:
            # Confirmed items always get recreated (user explicitly confirmed)
            if item_code in confirmed_set:
                can_create.append(node)
                continue

            # Already has active BOM - compare hashes using BOM's hash (source of truth)
            children = node.get("children", [])
            new_hash = calculate_tree_structure_hash(children)

            # Get existing hash from BOM itself
            existing_hash = frappe.db.get_value(
                "BOM", cm.active_bom, "custom_bom_structure_hash"
            )

            # Backfill if NULL
            if not existing_hash:
                bom_doc = frappe.get_doc("BOM", cm.active_bom)
                existing_hash = calculate_bom_doc_hash(bom_doc)
                if existing_hash:
                    frappe.db.set_value(
                        "BOM", cm.active_bom,
                        "custom_bom_structure_hash", existing_hash,
                        update_modified=False
                    )

            if new_hash == existing_hash:
                continue  # Unchanged, skip

        # Either new, confirmed, or needs BOM
        can_create.append(node)

    # Create BOMs
    analysis = {
        "can_create": can_create,
        "changed_components": [],
        "loose_blocked": [],
        "blocked_by_dependencies": {},
        "summary": {"can_create": len(can_create)},
    }

    # DEBUG: Log what we're about to create (consolidated log)
    debug_lines = [
        "=== BOM Upload Debug ===",
        f"confirmed_items: {confirmed_items}",
        f"can_create count: {len(can_create)}",
        f"can_create items: {[n['item_code'] for n in can_create]}",
        "",
        "=== Tree Structure ===",
    ]
    for node in all_assemblies:
        children_count = len(node.get("children", []))
        in_can_create = node["item_code"] in confirmed_set or any(n["item_code"] == node["item_code"] for n in can_create)
        debug_lines.append(f"  L{node.get('level')} | {node['item_code']} | children={children_count} | in_can_create={in_can_create}")

    frappe.log_error(
        title="DEBUG: BOM Upload Confirmation Flow",
        message="\n".join(debug_lines)
    )

    result = create_boms_and_link_components(tree, project, analysis, ws, image_loader, machine_code)

    # Build comprehensive summary (same format as create_boms_with_validation)
    # so _show_upload_success() can display the stats tables
    result["summary"] = {
        "items": {
            "created": item_counters["created"],
            "existing": item_counters["existing"],
            "updated": item_counters["updated"],
            "failed": item_counters["failed"],
            "total": sum(item_counters.values())
        },
        "boms": {
            "created": result.pop("created", 0),
            "existing": result.pop("skipped", 0),
            "failed": result.pop("failed", 0),
            "total": 0
        },
        "component_masters": {
            "created": 0,
            "existing": 0,
            "updated": 0,
            "failed": 0,
            "total": 0
        }
    }
    result["summary"]["boms"]["total"] = (
        result["summary"]["boms"]["created"] +
        result["summary"]["boms"]["existing"] +
        result["summary"]["boms"]["failed"]
    )

    # Count CMs for this project (they were created in the initial call)
    cm_count = frappe.db.count(
        "Project Component Master",
        {"project": project, "machine_code": machine_code}
    )
    result["summary"]["component_masters"]["existing"] = cm_count
    result["summary"]["component_masters"]["total"] = cm_count

    # Ensure status is set explicitly
    result["status"] = "success"

    result["message"] = _build_summary_message(result["summary"], result.get("errors", []))

    # Collect impacted parent BOMs — parents still referencing old child BOMs
    # via bom_no. Engineer needs to review these for controlled upward propagation.
    impacted = []
    for item_code, old_bom_name in old_boms.items():
        for parent in _get_impacted_parent_boms(old_bom_name):
            impacted.append({
                "changed_item": item_code,
                "old_bom": old_bom_name,
                "parent_bom": parent["parent_bom"],
                "parent_item": parent["parent_item"],
            })
    if impacted:
        result["impacted_parent_boms"] = impacted

    # Log upload history
    doc.append("upload_history", {
        "bom_file": doc.bom_file,
        "machine_code": machine_code,
        "uploaded_by": frappe.session.user,
        "uploaded_on": frappe.utils.now_datetime(),
        "status": "Success",
        "items_created": result["summary"]["items"].get("created", 0),
        "boms_created": result["summary"]["boms"].get("created", 0),
        "cms_created": result["summary"]["component_masters"].get("created", 0),
    })
    doc.save(ignore_permissions=True)

    frappe.log_error(title="BOM Upload Debug - Version Change", message=str(result))
    frappe.db.commit()
    return result


# ==================== Step 6 & 7: Create BOMs and Link ====================

def create_boms_and_link_components(tree, project, analysis, ws, image_loader, machine_code=None):
    """
    Create BOMs for components that passed analysis, then link active_bom
    back to their Component Masters.

    Args:
        tree: List of root nodes
        project: Project name
        analysis: Analysis results dict
        ws: Excel worksheet
        image_loader: SheetImageLoader instance (or None)
        machine_code: Machine code to scope hierarchy population (required for multi-machine)

    Returns:
        dict: {"status": "success", "created": X, "skipped": X, "failed": X, "errors": []}
    """
    created = 0
    skipped = 0
    failed = 0
    errors = []

    # Only create BOMs for items in the can_create list
    can_create_codes = {n["item_code"] for n in analysis["can_create"]}

    for node in tree:
        try:
            c, s, f, e = _create_bom_for_node(node, project, can_create_codes, ws, image_loader)
            created += c
            skipped += s
            failed += f
            errors.extend(e)
        except Exception as ex:
            failed += 1
            errors.append(f"{node['item_code']}: {str(ex)}")
            frappe.log_error(
                title=f"BOM creation failed for {node['item_code']}",
                message=frappe.get_traceback(),
            )

    # Link active_bom back to Component Masters
    _link_boms_to_component_masters(project)

    # Populate hierarchy codes (parent_component, m_code, g_code) for CMs with missing data
    # This runs after BOMs are linked so we have the full structure
    # machine_code filter prevents cross-machine contamination when same item exists in multiple machines
    _populate_hierarchy_codes(project, machine_code=machine_code, only_missing=True)

    # Refresh bom_usage rows' m_code/g_code from now-correct parent CM headers.
    # Must run AFTER _populate_hierarchy_codes because add_or_update_bom_usage() was
    # called before CM headers had m_code/g_code set (ordering issue).
    _refresh_bom_usage_hierarchy_codes(project, machine_code=machine_code)

    # Recalculate all Component Master quantities (synchronous, after commit)
    # This ensures total_qty_limit and other calculated fields are populated
    from clevertech.project_component_master.bom_hooks import recalculate_component_masters_for_project
    recalc_result = recalculate_component_masters_for_project(project)

    # Add recalculation info to errors if any failed
    if recalc_result.get("errors"):
        errors.extend(recalc_result["errors"])

    return {
        "status": "success",
        "created": created,
        "skipped": skipped,
        "failed": failed,
        "errors": errors,
    }


def _create_bom_for_node(node, project, can_create_codes, ws, image_loader):
    """
    Create BOM for a single node using existing create_bom_recursive.
    Only processes nodes whose item_code is in can_create_codes.

    Args:
        node: Tree node dict
        project: Project name
        can_create_codes: Set of item codes allowed to create BOMs
        ws: Excel worksheet
        image_loader: SheetImageLoader instance (or None)

    Returns:
        tuple: (created, skipped, failed, errors)
    """
    created = 0
    skipped = 0
    failed = 0
    errors = []

    item_code = node["item_code"]
    has_children = bool(node.get("children"))
    children_count = len(node.get("children", []))
    in_can_create = item_code in can_create_codes

    if not has_children:
        # Leaf node — no BOM needed
        return created, skipped, failed, errors

    if not in_can_create:
        skipped += 1
        # DON'T return here - still need to recurse into children!
    else:

        try:
            created_flag = create_bom_recursive(node, project, ws, image_loader)
            if created_flag:
                created += 1
            else:
                skipped += 1
        except Exception as ex:
            failed += 1
            errors.append(f"{item_code}: {str(ex)}")

    # ALWAYS recurse into children (even if current node was skipped)
    # This ensures we process confirmed child items even when parent is unchanged
    for child in node.get("children", []):
        if child.get("children"):
            c, s, f, e = _create_bom_for_node(child, project, can_create_codes, ws, image_loader)
            created += c
            skipped += s
            failed += f
            errors.extend(e)

    return created, skipped, failed, errors


def _link_boms_to_component_masters(project):
    """
    After BOMs are created, link active_bom back to Component Masters.
    Also populate BOM Usage for linked BOMs (since on_bom_submit doesn't
    trigger for existing BOMs from other projects).

    Fix (2026-02-04): Process ALL CMs with has_bom=1, not just those without
    active_bom. This handles the case where BOM is shared across projects -
    the BOM may have a different project set, so on_bom_submit wouldn't
    update this project's CM.

    Design: BOMs are shared across projects (no duplicates). When a version
    change is needed, a new BOM is created only for the affected project.
    """
    # Find ALL Component Masters that have BOMs (not just those without active_bom)
    # This ensures we update active_bom even if it was pointing to an old/wrong BOM
    cms = frappe.get_all(
        "Project Component Master",
        filters={
            "project": project,
            "has_bom": 1,
            # Removed "active_bom not set" filter - process ALL CMs with has_bom=1
        },
        fields=["name", "item_code", "active_bom"],  # Include current active_bom
    )

    # Track BOMs that were linked (not created) - need to populate BOM Usage for these
    linked_boms = []

    for cm_data in cms:
        # Find the active default BOM for this item (check with project first)
        bom_name = frappe.db.get_value(
            "BOM",
            {
                "item": cm_data.item_code,
                "project": project,
                "is_active": 1,
                "is_default": 1,
                "docstatus": 1,
            },
            "name",
        )

        if not bom_name:
            # Also check without project filter (BOM may belong to another project
            # but can be shared - this is the "copy-on-write" approach)
            bom_name = frappe.db.get_value(
                "BOM",
                {
                    "item": cm_data.item_code,
                    "is_active": 1,
                    "is_default": 1,
                    "docstatus": 1,
                },
                "name",
            )

        if bom_name:
            # Only update if active_bom is different (avoid unnecessary saves)
            if cm_data.active_bom == bom_name:
                continue  # Already pointing to correct BOM, skip

            cm = frappe.get_doc("Project Component Master", cm_data.name)
            cm.active_bom = bom_name

            # Recalculate hash from actual BOM doc
            try:
                bom_doc = frappe.get_doc("BOM", bom_name)
                cm.bom_structure_hash = calculate_bom_doc_hash(bom_doc)
            except Exception:
                pass

            cm.flags.ignore_validate = True
            cm.flags.ignore_mandatory = True
            cm.save(ignore_permissions=True)

            # Track this BOM for BOM Usage population
            linked_boms.append({"name": bom_name, "item": cm_data.item_code})

    # Populate BOM Usage for linked BOMs (since on_bom_submit didn't trigger)
    # This handles the case where BOMs already existed but Component Masters were just created
    if linked_boms:
        populate_bom_usage_tables(project, linked_boms)


def _populate_hierarchy_codes(project, machine_code=None, only_missing=True):
    """
    Populate parent_component, m_code, g_code for Component Masters in a project.

    This function is called after BOMs are linked to ensure hierarchy data is complete.
    It uses the BOM structure to derive:
    - parent_component: from active_bom's parent item
    - m_code: item starting with "M" in the hierarchy
    - g_code: item starting with "G" in the hierarchy

    Args:
        project: Project name
        machine_code: Machine code to filter by (prevents cross-machine contamination)
        only_missing: If True, only update CMs where fields are NULL (default)
                      If False, force refresh ALL CMs

    Returns:
        dict: {"updated": count, "skipped": count}
    """
    # Build filters based on only_missing flag
    # Include machine_code filter to prevent cross-machine contamination
    filters = {"project": project}
    if machine_code:
        filters["machine_code"] = machine_code

    if only_missing:
        # Get CMs where ANY of the hierarchy fields are missing
        sql_conditions = ["project = %(project)s"]
        sql_params = {"project": project}

        if machine_code:
            sql_conditions.append("machine_code = %(machine_code)s")
            sql_params["machine_code"] = machine_code

        cms = frappe.db.sql(f"""
            SELECT name, item_code, parent_component, m_code, g_code
            FROM `tabProject Component Master`
            WHERE {" AND ".join(sql_conditions)}
            AND (
                parent_component IS NULL OR parent_component = ''
                OR m_code IS NULL OR m_code = ''
                OR g_code IS NULL OR g_code = ''
            )
        """, sql_params, as_dict=True)
    else:
        cms = frappe.get_all(
            "Project Component Master",
            filters=filters,
            fields=["name", "item_code", "parent_component", "m_code", "g_code"]
        )

    if not cms:
        return {"updated": 0, "skipped": 0}

    updated = 0
    skipped = 0

    # Build a lookup of item_code → CM name for parent lookups
    # Filter by machine_code to prevent cross-machine contamination
    all_cms = frappe.get_all(
        "Project Component Master",
        filters=filters,
        fields=["name", "item_code"]
    )
    item_to_cm = {cm.item_code: cm.name for cm in all_cms}

    for cm_data in cms:
        item_code = cm_data.item_code
        has_changes = False

        # Derive m_code and g_code from item_code itself
        derived_m_code = None
        derived_g_code = None

        if item_code.startswith("M"):
            # This IS an M-code
            derived_m_code = item_code
            derived_g_code = None
        elif item_code.startswith("G"):
            # This IS a G-code, need to find its M-code from parent
            derived_g_code = item_code
            # Find parent via BOM structure
            parent_item = _find_parent_item_via_bom(item_code, project)
            if parent_item and parent_item.startswith("M"):
                derived_m_code = parent_item
            elif parent_item:
                # Parent is not M-code, try to get m_code from parent CM
                parent_cm_name = item_to_cm.get(parent_item)
                if parent_cm_name:
                    derived_m_code = frappe.db.get_value(
                        "Project Component Master", parent_cm_name, "m_code"
                    )
        else:
            # This is below G-level (D-code, raw material, etc.)
            # Find parent and derive both m_code and g_code
            parent_item = _find_parent_item_via_bom(item_code, project)
            if parent_item:
                parent_cm_name = item_to_cm.get(parent_item)
                if parent_cm_name:
                    parent_cm_data = frappe.db.get_value(
                        "Project Component Master",
                        parent_cm_name,
                        ["m_code", "g_code", "item_code"],
                        as_dict=True
                    )
                    if parent_cm_data:
                        # Inherit m_code from parent
                        derived_m_code = parent_cm_data.m_code

                        # If parent is G-code, use it as g_code
                        if parent_cm_data.item_code.startswith("G"):
                            derived_g_code = parent_cm_data.item_code
                        else:
                            # Otherwise inherit parent's g_code
                            derived_g_code = parent_cm_data.g_code

        # Determine parent_component
        derived_parent_component = None
        parent_item = _find_parent_item_via_bom(item_code, project)
        if parent_item:
            derived_parent_component = item_to_cm.get(parent_item)

        # Check what needs updating
        needs_parent = (not cm_data.parent_component) and derived_parent_component
        needs_m_code = (not cm_data.m_code) and derived_m_code
        needs_g_code = (not cm_data.g_code) and derived_g_code

        if needs_parent or needs_m_code or needs_g_code or not only_missing:
            # Update using db_set for efficiency (avoid full doc load/save)
            updates = {}
            if needs_parent or (not only_missing and derived_parent_component):
                updates["parent_component"] = derived_parent_component
            if needs_m_code or (not only_missing and derived_m_code):
                updates["m_code"] = derived_m_code
            if needs_g_code or (not only_missing and derived_g_code):
                updates["g_code"] = derived_g_code

            if updates:
                frappe.db.set_value(
                    "Project Component Master",
                    cm_data.name,
                    updates,
                    update_modified=False
                )
                has_changes = True

        if has_changes:
            updated += 1
        else:
            skipped += 1

    if updated > 0:
        frappe.db.commit()

    return {"updated": updated, "skipped": skipped}


def _find_parent_item_via_bom(item_code, project):
    """
    Find the parent item of an item by looking at BOM structure.

    Checks which BOMs contain this item as a child and returns the parent item.
    Prioritizes BOMs with the same project, then any active BOM.

    Args:
        item_code: Item code to find parent for
        project: Project name

    Returns:
        str: Parent item code, or None if not found
    """
    # First try: Find BOM with matching project that contains this item
    parent_item = frappe.db.sql("""
        SELECT DISTINCT bom.item
        FROM `tabBOM` bom
        INNER JOIN `tabBOM Item` bi ON bi.parent = bom.name
        WHERE bi.item_code = %(item_code)s
        AND bom.project = %(project)s
        AND bom.docstatus = 1
        AND bom.is_active = 1
        LIMIT 1
    """, {"item_code": item_code, "project": project})

    if parent_item:
        return parent_item[0][0]

    # Second try: Any active BOM that contains this item
    parent_item = frappe.db.sql("""
        SELECT DISTINCT bom.item
        FROM `tabBOM` bom
        INNER JOIN `tabBOM Item` bi ON bi.parent = bom.name
        WHERE bi.item_code = %(item_code)s
        AND bom.docstatus = 1
        AND bom.is_active = 1
        LIMIT 1
    """, {"item_code": item_code})

    if parent_item:
        return parent_item[0][0]

    return None


def _refresh_bom_usage_hierarchy_codes(project, machine_code=None):
    """
    Refresh m_code and g_code on all bom_usage rows for a project.

    This fixes the ordering issue where add_or_update_bom_usage() was called
    BEFORE _populate_hierarchy_codes() set the CM header m_code/g_code.
    At that point, parent CMs had NULL m_code, so _derive_codes_from_parent()
    returned NULL for items under G-codes.

    Now that CM headers are correct, re-derive bom_usage m_code/g_code from
    the parent item code and parent CM's header fields.

    Args:
        project: Project name
        machine_code: Machine code to filter by (prevents cross-machine contamination)

    Returns:
        dict: {"updated": count, "skipped": count}
    """
    # Get all bom_usage rows with missing m_code or g_code
    rows = frappe.db.sql("""
        SELECT
            cbu.name AS row_name,
            cbu.parent AS cm_name,
            cbu.parent_bom,
            cbu.m_code AS current_m_code,
            cbu.g_code AS current_g_code,
            pcm.item_code
        FROM `tabComponent BOM Usage` cbu
        INNER JOIN `tabProject Component Master` pcm ON pcm.name = cbu.parent
        WHERE pcm.project = %(project)s
        AND (cbu.m_code IS NULL OR cbu.m_code = ''
             OR cbu.g_code IS NULL OR cbu.g_code = '')
    """, {"project": project}, as_dict=True)

    if not rows:
        return {"updated": 0, "skipped": 0}

    # Build lookup: item_code → CM data (m_code, g_code)
    # Filter by machine_code to prevent cross-machine contamination
    filters = {"project": project}
    if machine_code:
        filters["machine_code"] = machine_code

    all_cms = frappe.get_all(
        "Project Component Master",
        filters=filters,
        fields=["name", "item_code", "m_code", "g_code"]
    )
    item_to_cm = {cm.item_code: cm for cm in all_cms}

    updated = 0
    skipped = 0

    for row in rows:
        # Get the parent item from the parent BOM
        parent_item = None
        if row.parent_bom:
            parent_item = frappe.db.get_value("BOM", row.parent_bom, "item")

        if not parent_item:
            skipped += 1
            continue

        # Derive m_code/g_code from parent item (same logic as _derive_codes_from_parent)
        derived_m = None
        derived_g = None

        if parent_item.startswith("M"):
            derived_m = parent_item
            derived_g = None
        elif parent_item.startswith("G"):
            derived_g = parent_item
            # Get m_code from parent CM's header (now populated by _populate_hierarchy_codes)
            parent_cm_data = item_to_cm.get(parent_item)
            if parent_cm_data:
                derived_m = parent_cm_data.m_code
        else:
            # Parent is D-code or deeper — inherit from parent CM header
            parent_cm_data = item_to_cm.get(parent_item)
            if parent_cm_data:
                derived_m = parent_cm_data.m_code
                if parent_cm_data.item_code.startswith("G"):
                    derived_g = parent_cm_data.item_code
                else:
                    derived_g = parent_cm_data.g_code

        # Only update if we have something better than current values
        needs_update = False
        update_fields = {}

        if not row.current_m_code and derived_m:
            update_fields["m_code"] = derived_m
            needs_update = True
        if not row.current_g_code and derived_g:
            update_fields["g_code"] = derived_g
            needs_update = True

        if needs_update:
            frappe.db.set_value(
                "Component BOM Usage", row.row_name,
                update_fields, update_modified=False
            )
            updated += 1
        else:
            skipped += 1

    if updated > 0:
        frappe.db.commit()

    return {"updated": updated, "skipped": skipped}


# ==================== Hash Calculation ====================

def calculate_tree_structure_hash(children):
    """
    Create hash of BOM structure from tree node children (parsed from Excel).
    Uses same logic as bom_hooks.calculate_bom_structure_hash but works
    on tree node dicts instead of BOM doc items.

    Args:
        children: List of child node dicts with item_code and qty

    Returns:
        str: MD5 hash of sorted (item_code, qty) tuples
    """
    if not children:
        return None

    structure = sorted(
        [(child["item_code"], float(child.get("qty", 1))) for child in children],
        key=lambda x: x[0],
    )
    structure_str = json.dumps(structure, sort_keys=True)
    return hashlib.md5(structure_str.encode()).hexdigest()


def calculate_bom_diff(old_bom_name, new_children):
    """
    Calculate diff between old BOM and new tree children.
    Shows what items were added, removed, or had qty changes.

    Args:
        old_bom_name: Name of the old BOM
        new_children: List of child dicts from parsed Excel tree

    Returns:
        dict: {
            "added": [{"item_code": str, "qty": float}, ...],
            "removed": [{"item_code": str, "qty": float}, ...],
            "qty_changed": [{"item_code": str, "old_qty": float, "new_qty": float}, ...]
        }
    """
    # Get old BOM items (consolidated)
    old_items = {}
    if old_bom_name:
        try:
            old_bom_doc = frappe.get_doc("BOM", old_bom_name)
            for item in old_bom_doc.items:
                if item.item_code in old_items:
                    old_items[item.item_code] += float(item.qty or 0)
                else:
                    old_items[item.item_code] = float(item.qty or 0)
        except frappe.DoesNotExistError:
            pass

    # Get new items from tree children (consolidated)
    new_items = {}
    for child in (new_children or []):
        item_code = child.get("item_code")
        qty = float(child.get("qty", 0))
        if item_code in new_items:
            new_items[item_code] += qty
        else:
            new_items[item_code] = qty

    old_set = set(old_items.keys())
    new_set = set(new_items.keys())

    added = [{"item_code": ic, "qty": new_items[ic]} for ic in sorted(new_set - old_set)]
    removed = [{"item_code": ic, "qty": old_items[ic]} for ic in sorted(old_set - new_set)]
    qty_changed = []

    for ic in sorted(old_set & new_set):
        if abs(old_items[ic] - new_items[ic]) > 0.0001:  # Float comparison tolerance
            qty_changed.append({
                "item_code": ic,
                "old_qty": old_items[ic],
                "new_qty": new_items[ic]
            })

    return {
        "added": added,
        "removed": removed,
        "qty_changed": qty_changed
    }


# ==================== Summary Message ====================

def _build_summary_message(summary, errors=None):
    """
    Build a human-readable summary message.

    Args:
        summary: dict with items, boms, component_masters sections
        errors: list of error strings

    Returns:
        str: Formatted summary message
    """
    parts = []

    items = summary.get("items", {})
    if items.get("created"):
        parts.append(f"{items['created']} Items created")
    if items.get("existing"):
        parts.append(f"{items['existing']} Items already existed")

    boms = summary.get("boms", {})
    if boms.get("created"):
        parts.append(f"{boms['created']} BOMs created")
    if boms.get("existing"):
        parts.append(f"{boms['existing']} BOMs already existed")

    cms = summary.get("component_masters", {})
    if cms.get("created"):
        parts.append(f"{cms['created']} Component Masters created")
    if cms.get("existing"):
        parts.append(f"{cms['existing']} Component Masters already existed")
    if cms.get("updated"):
        parts.append(f"{cms['updated']} Component Masters updated")

    # Add failure warnings
    failures = items.get("failed", 0) + boms.get("failed", 0) + cms.get("failed", 0)
    if failures:
        parts.append(f"⚠️ {failures} failures")

    if errors and len(errors) > 0:
        parts.append(f"⚠️ {len(errors)} errors (check error log)")

    return ". ".join(parts) + "." if parts else "No changes made."


# ==================== Serialization ====================

def _serialize_analysis(analysis):
    """
    Serialize analysis dict for JSON response to client.
    Convert node dicts to safe serializable format.
    """
    def serialize_node(node):
        return {
            "item_code": node.get("item_code"),
            "description": node.get("description"),
            "level": node.get("level"),
            "qty": node.get("qty"),
        }

    def serialize_details(details):
        """Serialize details dict, ensuring procurement_docs is serializable."""
        if not details:
            return {}
        return {
            "change_type": details.get("change_type"),
            "component_master": details.get("component_master"),
            "old_bom": details.get("old_bom"),
            "blocking_level": details.get("blocking_level"),
            "blocking_message": details.get("blocking_message"),
            "can_proceed": details.get("can_proceed"),
            "procurement_docs": details.get("procurement_docs", []),
            "bom_diff": details.get("bom_diff", {}),
            "impacted_parents": details.get("impacted_parents", []),
        }

    serialized = {
        "can_create": [serialize_node(n) for n in analysis.get("can_create", [])],
        "changed_components": [],
        "loose_blocked": [],
        "blocked_by_dependencies": analysis.get("blocked_by_dependencies", {}),
        "summary": analysis.get("summary", {}),
    }

    for comp in analysis.get("changed_components", []):
        serialized["changed_components"].append({
            "node": serialize_node(comp.get("node", {})),
            "details": serialize_details(comp.get("details", {})),
        })

    for comp in analysis.get("loose_blocked", []):
        serialized["loose_blocked"].append({
            "node": serialize_node(comp.get("node", {})),
            "details": comp.get("details", {}),
        })

    return serialized


# ==================== Debug Functions ====================

@frappe.whitelist()
def debug_bom_quantities(docname, target_item_code=None):
    """
    Debug function to trace BOM quantity calculations without creating BOMs.

    Simulates the BOM creation process and reports how quantities are calculated
    for a specific item code, helping identify discrepancies between Excel and BOM.

    Args:
        docname: BOM Upload document name
        target_item_code: Specific item code to trace (optional, if None traces all)

    Returns:
        dict: {
            "target_item": str,
            "excel_total_qty": float,
            "excel_occurrences": int,
            "occurrences_by_parent": [
                {
                    "parent_item": str,
                    "parent_level": int,
                    "occurrences": [
                        {
                            "row": int,
                            "position": str,
                            "qty": float,
                            "level": int
                        }
                    ],
                    "total_qty": float,
                    "bom_items_count": int  # How many BOM items would be created
                }
            ],
            "tree_analysis": str  # Human-readable analysis
        }
    """
    doc = frappe.get_doc("BOM Upload", docname)

    if not doc.bom_file:
        frappe.throw(_("Please attach a BOM Excel file first."))

    # Step 1: Parse Excel
    file_doc = frappe.get_doc("File", {"file_url": doc.bom_file})
    content = file_doc.get_content()

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = parse_rows_dynamic(ws)
    tree = build_tree(rows)

    if not tree:
        frappe.throw(_("No components found in the Excel file."))

    # Step 2: If no target specified, find most common item for demo
    if not target_item_code:
        # Count item occurrences
        item_counts = {}
        def count_items(nodes):
            for node in nodes:
                item_counts[node["item_code"]] = item_counts.get(node["item_code"], 0) + 1
                if node["children"]:
                    count_items(node["children"])
        count_items(tree)

        # Find item that appears most frequently
        if item_counts:
            target_item_code = max(item_counts.items(), key=lambda x: x[1])[0]
        else:
            frappe.throw(_("No items found in Excel."))

    # Step 3: Trace all occurrences of target item
    occurrences_by_parent = {}
    excel_total_qty = 0
    excel_occurrences_count = 0

    def trace_item_in_tree(parent_node, parent_stack):
        """
        Recursively trace occurrences of target item.

        Args:
            parent_node: Current node being examined
            parent_stack: List of parent nodes leading to this node
        """
        nonlocal excel_total_qty, excel_occurrences_count

        # Check if any direct children match target item
        for child in parent_node["children"]:
            if child["item_code"] == target_item_code:
                # Found an occurrence
                excel_total_qty += child["qty"]
                excel_occurrences_count += 1

                # Track by parent assembly
                parent_key = parent_node["item_code"]
                if parent_key not in occurrences_by_parent:
                    occurrences_by_parent[parent_key] = {
                        "parent_item": parent_node["item_code"],
                        "parent_description": parent_node.get("description", ""),
                        "parent_level": parent_node["level"],
                        "occurrences": [],
                        "total_qty": 0
                    }

                occurrences_by_parent[parent_key]["occurrences"].append({
                    "row": child.get("row_num", "?"),
                    "position": child.get("position", "?"),
                    "qty": child["qty"],
                    "level": child["level"],
                    "parent_chain": " → ".join([p["item_code"] for p in parent_stack] + [parent_node["item_code"]])
                })
                occurrences_by_parent[parent_key]["total_qty"] += child["qty"]

            # Recurse into child if it has children (is an assembly)
            if child["children"]:
                trace_item_in_tree(child, parent_stack + [parent_node])

    # Start tracing from all root nodes
    for root in tree:
        if root["children"]:
            trace_item_in_tree(root, [])

    # Step 4: Calculate how many BOM items would be created per parent
    for parent_key, data in occurrences_by_parent.items():
        # In current implementation, each occurrence creates a separate BOM item
        # (no consolidation happens in build_tree or create_bom_recursive)
        data["bom_items_count"] = len(data["occurrences"])

    # Step 5: Generate human-readable analysis
    analysis_lines = []
    analysis_lines.append(f"=== BOM Quantity Debug Report ===")
    analysis_lines.append(f"Target Item: {target_item_code}")
    analysis_lines.append(f"")
    analysis_lines.append(f"Excel Analysis:")
    analysis_lines.append(f"  Total Occurrences: {excel_occurrences_count}")
    analysis_lines.append(f"  Total Quantity (sum): {excel_total_qty}")
    analysis_lines.append(f"")
    analysis_lines.append(f"Occurrences by Parent Assembly:")
    analysis_lines.append(f"")

    for parent_key in sorted(occurrences_by_parent.keys()):
        data = occurrences_by_parent[parent_key]
        analysis_lines.append(f"Parent: {data['parent_item']} (Level {data['parent_level']})")
        analysis_lines.append(f"  Description: {data['parent_description']}")
        analysis_lines.append(f"  Occurrences: {len(data['occurrences'])}")
        analysis_lines.append(f"  Total Qty: {data['total_qty']}")
        analysis_lines.append(f"  BOM Items to be Created: {data['bom_items_count']}")
        analysis_lines.append(f"")

        for i, occ in enumerate(data["occurrences"], 1):
            analysis_lines.append(f"    #{i}: Row {occ['row']}, Pos={occ['position']}, Qty={occ['qty']}, Level={occ['level']}")
            analysis_lines.append(f"        Hierarchy: {occ['parent_chain']}")

        analysis_lines.append(f"")

        # CRITICAL: Explain how BOM creation works
        if data['bom_items_count'] > 1:
            analysis_lines.append(f"  ⚠️  IMPORTANT: {data['parent_item']} will have {data['bom_items_count']} SEPARATE")
            analysis_lines.append(f"      BOM item entries for {target_item_code} (one per Excel row).")
            analysis_lines.append(f"      Frappe BOM may or may not consolidate these - need to check actual BOM!")
            analysis_lines.append(f"")

    analysis_lines.append(f"=== How BOM Creation Works ===")
    analysis_lines.append(f"1. build_tree() appends each Excel row to parent's children[] (NO consolidation)")
    analysis_lines.append(f"2. create_bom_recursive() iterates children[] and adds each as BOM item")
    analysis_lines.append(f"3. If same item appears 3 times in Excel → 3 entries in children[] → 3 BOM items")
    analysis_lines.append(f"4. Frappe may consolidate duplicate item_code entries in BOM UI (TBD - check actual BOM)")
    analysis_lines.append(f"")

    # Convert to list for JSON serialization
    occurrences_list = [data for data in occurrences_by_parent.values()]

    return {
        "target_item": target_item_code,
        "excel_total_qty": excel_total_qty,
        "excel_occurrences": excel_occurrences_count,
        "occurrences_by_parent": occurrences_list,
        "tree_analysis": "\n".join(analysis_lines),
        "parents_count": len(occurrences_by_parent)
    }


# ==================== Debug: Full Upload Flow Trace ====================

@frappe.whitelist()
def debug_upload_flow(docname):
    """
    Trace the upload flow step by step using the same Excel file and project.
    Does NOT create/modify anything — only reads current DB state and logs
    what each stage would see / did produce.
    """
    log = []  # list of {stage, message}

    def add(stage, msg):
        log.append({"stage": stage, "message": msg})

    # --- Load doc, parse Excel (same as create_boms_with_validation) ---
    doc = frappe.get_doc("BOM Upload", docname)
    project = doc.project
    machine_code = doc.machine_code
    add("setup", f"Project: {project}")
    add("setup", f"Machine Code: {machine_code}")

    file_doc = frappe.get_doc("File", {"file_url": doc.bom_file})
    content = file_doc.get_content()
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = parse_rows_dynamic(ws)
    tree = build_tree(rows)
    all_nodes = _get_all_nodes(tree)
    add("parse", f"Excel parsed: {len(rows)} rows → {len(tree)} root(s), {len(all_nodes)} total nodes")

    # Log tree structure
    for node in all_nodes:
        children_count = len(node.get("children", []))
        add("parse", f"  L{node.get('level')} | {node['item_code']} | qty={node.get('qty')} | children={children_count}")

    # --- Stage: Component Masters ---
    add("cm", "--- Checking Component Masters ---")
    for node in all_nodes:
        item_code = node["item_code"]
        level = node.get("level", 0)
        cm_name = frappe.db.get_value(
            "Project Component Master",
            {"project": project, "item_code": item_code, "machine_code": machine_code},
            "name"
        )
        if cm_name:
            cm = frappe.get_doc("Project Component Master", cm_name)
            add("cm", (
                f"  {item_code} → EXISTS ({cm_name}) | "
                f"make_or_buy={cm.make_or_buy} | "
                f"project_qty={cm.project_qty} | "
                f"bom_qty_required={cm.bom_qty_required} | "
                f"total_qty_limit={cm.total_qty_limit} | "
                f"has_bom={cm.has_bom} | active_bom={cm.active_bom}"
            ))
        else:
            add("cm", f"  {item_code} → NOT FOUND (would be created)")

    # --- Stage: BOMs ---
    add("bom", "--- Checking BOMs ---")
    for node in all_nodes:
        if not node.get("children"):
            continue  # leaf — no BOM
        item_code = node["item_code"]
        bom_name = frappe.db.get_value("BOM", {
            "item": item_code, "is_active": 1, "is_default": 1, "docstatus": 1
        }, "name")
        if bom_name:
            bom_doc = frappe.get_doc("BOM", bom_name)
            child_codes = [i.item_code for i in bom_doc.items]
            add("bom", f"  {item_code} → BOM exists: {bom_name} | project={bom_doc.project} | items={child_codes}")
        else:
            add("bom", f"  {item_code} → NO BOM (would be created)")

    # --- Stage: BOM Hash Comparison ---
    add("hash", "--- BOM Hash Comparison (Excel vs BOM vs CM) ---")
    for node in all_nodes:
        if not node.get("children"):
            continue  # leaf — no BOM, no hash
        item_code = node["item_code"]

        # Excel hash — computed from this upload's tree
        excel_hash = calculate_tree_structure_hash(node["children"])

        # BOM hash — stored on the submitted BOM
        bom_name = frappe.db.get_value("BOM", {
            "item": item_code, "is_active": 1, "is_default": 1, "docstatus": 1
        }, "name")
        bom_hash = None
        if bom_name:
            bom_hash = frappe.db.get_value("BOM", bom_name, "custom_bom_structure_hash")

        # CM hash — stored on the Component Master
        cm_name = frappe.db.get_value(
            "Project Component Master",
            {"project": project, "item_code": item_code, "machine_code": machine_code},
            "name"
        )
        cm_hash = None
        if cm_name:
            cm_hash = frappe.db.get_value("Project Component Master", cm_name, "bom_structure_hash")

        # Compare
        excel_vs_bom = "HASH MATCH" if (excel_hash == bom_hash) else "HASH MISMATCH"
        excel_vs_cm  = "HASH MATCH" if (excel_hash == cm_hash)  else "HASH MISMATCH"

        add("hash", (
            f"  {item_code} | "
            f"Excel={excel_hash} | "
            f"BOM={bom_hash} ({excel_vs_bom}) | "
            f"CM={cm_hash} ({excel_vs_cm})"
        ))

    # --- Stage: BOM Usage tables ---
    add("usage", "--- BOM Usage on each CM ---")
    for node in all_nodes:
        item_code = node["item_code"]
        cm_name = frappe.db.get_value(
            "Project Component Master",
            {"project": project, "item_code": item_code, "machine_code": machine_code},
            "name"
        )
        if not cm_name:
            continue
        cm = frappe.get_doc("Project Component Master", cm_name)
        if not cm.bom_usage:
            add("usage", f"  {item_code} → bom_usage: EMPTY")
        else:
            for row in cm.bom_usage:
                add("usage", (
                    f"  {item_code} → parent_bom={row.parent_bom} | "
                    f"parent_item={row.parent_item} | "
                    f"parent_component={row.parent_component} | "
                    f"qty_per_unit={row.qty_per_unit} | "
                    f"m_code={row.m_code} | g_code={row.g_code}"
                ))

    # --- Stage: Recalculation (step by step, in level order) ---
    add("recalc", "--- Recalculation (mirrors calculate_bom_qty_required) ---")
    for node in all_nodes:
        item_code = node["item_code"]
        cm_name = frappe.db.get_value(
            "Project Component Master",
            {"project": project, "item_code": item_code, "machine_code": machine_code},
            "name"
        )
        if not cm_name:
            continue
        cm = frappe.get_doc("Project Component Master", cm_name)

        if not cm.bom_usage:
            if cm.is_loose_item:
                add("recalc", f"  {item_code} → no bom_usage + loose → bom_qty_required = 0")
            else:
                add("recalc", f"  {item_code} → no bom_usage (root) → bom_qty_required = project_qty = {cm.project_qty}")
        else:
            total = 0
            for usage in cm.bom_usage:
                parent = frappe.db.get_value(
                    "Project Component Master",
                    {"project": cm.project, "item_code": usage.parent_item},
                    ["total_qty_limit", "make_or_buy", "name"],
                    as_dict=True
                )
                if not parent:
                    add("recalc", (
                        f"  {item_code} → PARENT LOOKUP FAILED for parent_item={usage.parent_item} "
                        f"(project={cm.project})"
                    ))
                elif parent.make_or_buy != "Make":
                    add("recalc", (
                        f"  {item_code} → parent {usage.parent_item} ({parent.name}) "
                        f"is {parent.make_or_buy} → skipped"
                    ))
                else:
                    row_total = float(usage.qty_per_unit or 0) * float(parent.total_qty_limit or 0)
                    total += row_total
                    add("recalc", (
                        f"  {item_code} → {usage.qty_per_unit} × parent.total_qty_limit({parent.total_qty_limit}) "
                        f"= {row_total}  [parent={usage.parent_item} / {parent.name}]"
                    ))
            tql = max(float(cm.project_qty or 0), total)
            add("recalc", (
                f"  {item_code} → RESULT: bom_qty_required={total} | "
                f"total_qty_limit=MAX({cm.project_qty},{total})={tql} | "
                f"DB has: bom_qty={cm.bom_qty_required}, tql={cm.total_qty_limit}"
            ))

    return log
