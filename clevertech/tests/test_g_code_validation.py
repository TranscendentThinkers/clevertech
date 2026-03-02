"""
Integration Test Suite for BOM Upload → Procurement Validation

Tests the full flow: Excel upload → BOM/CM creation → MR validation.
Uses real bom_upload_enhanced.py to create test data (not manual CM creation).

Test Scenarios:
  1. G-code level MR validation (Decision 20)
  2. Make/Buy enforcement
  3. Cross-machine CM isolation
  4. Cumulative MR qty tracking

Run:
    bench --site clevertech-uat.bharatbodh.com run-tests \
        --app clevertech --module clevertech.tests.test_g_code_validation
"""

import frappe
import openpyxl
import io
from frappe.tests.utils import FrappeTestCase
from frappe.utils.file_manager import save_file


# ============================================================================
# Excel Helper — creates minimal BOM Excel that bom_upload_enhanced can parse
# ============================================================================

def create_test_excel(rows):
    """
    Create a BOM Excel file in memory.

    Args:
        rows: list of dicts with keys:
            item_code, description, qty, level, state (optional)

    Returns:
        bytes: Excel file content
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    # Row 1: Title
    ws["A1"] = "Test BOM Upload"

    # Row 2: Headers (must match bom_upload_enhanced.py dynamic column mapping)
    headers = {
        "A": "Position",
        "C": "Item no",
        "D": "Description",
        "E": "Qty",
        "F": "Rev.",
        "G": "DESCRIZIONE_ESTESA",
        "H": "MATERIAL",
        "I": "Part_number",
        # J = STATE (hardcoded read in parse_rows_dynamic, no header needed)
        "K": "WEIGHT",
        "L": "MANUFACTURER",
        "M": "TIPO_TRATTAMENTO",
        "N": "UM",
        "O": "LivelloBom",
    }
    for col, header in headers.items():
        ws[f"{col}2"] = header

    # Row 3+: Data
    for i, row in enumerate(rows):
        r = i + 3  # Start at row 3
        ws[f"A{r}"] = i + 1                          # Position
        ws[f"C{r}"] = row["item_code"]                # Item no
        ws[f"D{r}"] = row.get("description", "")      # Description
        ws[f"E{r}"] = row.get("qty", 1)               # Qty
        ws[f"F{r}"] = row.get("revision", "01")        # Rev.
        ws[f"G{r}"] = row.get("description", "")      # Extended description
        ws[f"J{r}"] = row.get("state", "")             # STATE (column J, hardcoded)
        ws[f"N{r}"] = row.get("uom", "Nos")           # UM
        ws[f"O{r}"] = row["level"]                     # LivelloBom

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def run_bom_upload(project_name, machine_code, excel_rows, upload_doc_name=None):
    """
    Create or reuse BOM Upload doc, attach Excel, and run the full upload flow.

    BOM Upload autoname = "field:project", so only one doc per project.
    For second machine upload, reuse the same doc (update machine_code + file).

    Returns:
        tuple: (result dict, upload doc name)
    """
    from clevertech.clevertech.doctype.bom_upload.bom_upload_enhanced import (
        create_boms_with_validation,
    )

    if upload_doc_name and frappe.db.exists("BOM Upload", upload_doc_name):
        # Reuse existing doc (second machine upload on same project)
        upload_doc = frappe.get_doc("BOM Upload", upload_doc_name)
        upload_doc.machine_code = machine_code
    else:
        # Create new BOM Upload document
        upload_doc = frappe.get_doc({
            "doctype": "BOM Upload",
            "project": project_name,
            "machine_code": machine_code,
        }).insert(ignore_permissions=True)

    # Create and attach Excel file
    excel_bytes = create_test_excel(excel_rows)
    file_doc = save_file(
        f"test_bom_{machine_code}.xlsx",
        excel_bytes,
        "BOM Upload",
        upload_doc.name,
        is_private=1,
    )
    upload_doc.bom_file = file_doc.file_url
    upload_doc.save(ignore_permissions=True)
    frappe.db.commit()

    # Run the upload
    result = create_boms_with_validation(upload_doc.name)
    frappe.db.commit()

    return result, upload_doc.name


# ============================================================================
# Test Data Definitions
# ============================================================================

# Machine 1: MT4000084237
# Hierarchy (realistic: M → G → D → RM):
#   MT4000084237 (level 0, Machine) — Make
#   ├── GT3000012345 (level 1, G-code 1, RELEASED) — Make
#   │   └── DT0000054321 (level 2, D-code sub-assembly) — Buy (default)
#   │       └── RM0000012345 (level 3, Raw Material, qty=60) — Buy
#   ├── GT3000067890 (level 1, G-code 2, RELEASED) — Make
#   │   └── DT0000054322 (level 2, D-code sub-assembly) — Buy (default)
#   │       └── RM0000012345 (level 3, same RM, qty=40) — Buy
#   └── RM0000099999 (level 1, RM directly under M, qty=5) — Buy (tests RM under Make parent)
#
# After upload (with D-codes as Buy):
#   RM0000012345 CM: total_qty_limit = 0 (parent D-codes are Buy → RM not needed)
#   DT0000054321 CM: total_qty_limit > 0 (Buy → can be procured)
#
# After changing D-codes to Make:
#   RM0000012345 CM: total_qty_limit = 100 (60 + 40, recalculated from BOM explosion)

MACHINE_1_CODE = "VT0000000001"  # Machine code (not item code)
MACHINE_1_ITEM = "MT4000084237"
MACHINE_1_ROWS = [
    {"item_code": "MT4000084237",  "description": "Test Machine Assembly 1",    "qty": 1,  "level": 0, "state": "RELEASED"},
    {"item_code": "GT3000012345",  "description": "Test G-Code Assembly 1",     "qty": 1,  "level": 1, "state": "RELEASED"},
    {"item_code": "DT0000054321",  "description": "Test D-Code Sub-Assembly 1", "qty": 1,  "level": 2},
    {"item_code": "AT0000012345",  "description": "Test Raw Material Screw",    "qty": 60, "level": 3},
    {"item_code": "GT3000067890",  "description": "Test G-Code Assembly 2",     "qty": 1,  "level": 1, "state": "RELEASED"},
    {"item_code": "DT0000054322",  "description": "Test D-Code Sub-Assembly 2", "qty": 1,  "level": 2},
    {"item_code": "AT0000012345",  "description": "Test Raw Material Screw",    "qty": 40, "level": 3},
    {"item_code": "AT0000099999",  "description": "Test Raw Material Bolt",     "qty": 5,  "level": 1},
]

# Machine 2: MT4000084238 (same RM, different machine — tests CM isolation)
# Hierarchy:
#   MT4000084238 (level 0, Machine) — Make
#   └── GT3000099999 (level 1, G-code 3, RELEASED) — Make
#       └── DT0000054323 (level 2, D-code) — Buy (default)
#           └── RM0000012345 (level 3, same RM, qty=20) — Buy
#
# After upload:
#   RM0000012345 gets a SECOND CM (machine_code=MT4000084238)

MACHINE_2_CODE = "VT0000000002"  # Machine code (not item code)
MACHINE_2_ITEM = "MT4000084238"
MACHINE_2_ROWS = [
    {"item_code": "MT4000084238",  "description": "Test Machine Assembly 2",    "qty": 1,  "level": 0, "state": "RELEASED"},
    {"item_code": "GT3000099999",  "description": "Test G-Code Assembly 3",     "qty": 1,  "level": 1, "state": "RELEASED"},
    {"item_code": "DT0000054323",  "description": "Test D-Code Sub-Assembly 3", "qty": 1,  "level": 2},
    {"item_code": "AT0000012345",  "description": "Test Raw Material Screw",    "qty": 20, "level": 3},
]

RAW_MATERIAL = "AT0000012345"  # Using A prefix which has denomination mapping
RAW_MATERIAL_2 = "AT0000099999"
D_CODE_1 = "DT0000054321"
D_CODE_2 = "DT0000054322"
D_CODE_3 = "DT0000054323"
G_CODE_1 = "GT3000012345"
G_CODE_2 = "GT3000067890"
G_CODE_3 = "GT3000099999"


# ============================================================================
# Test Class
# ============================================================================

class TestGCodeValidation(FrappeTestCase):
    """Integration tests: BOM Upload → MR/PO Validation"""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        frappe.set_user("Administrator")

        # --- Project ---
        cls.project = frappe.get_doc({
            "doctype": "Project",
            "project_name": "TEST-G-CODE-VALIDATION",
            "status": "Open",
        }).insert(ignore_permissions=True)

        # --- Cost Centers (with custom_machine_code for MR filtering) ---
        cls.company = frappe.defaults.get_defaults().company
        root_cc = frappe.db.get_value(
            "Cost Center",
            {"company": cls.company, "is_group": 1},
            "name",
            order_by="lft asc",
        )

        cls.cost_center_1 = frappe.get_doc({
            "doctype": "Cost Center",
            "cost_center_name": "Test MT4000084237",
            "company": cls.company,
            "parent_cost_center": root_cc,
            "is_group": 0,
            "custom_machine_code": MACHINE_1_CODE,
        }).insert(ignore_permissions=True)

        cls.cost_center_2 = frappe.get_doc({
            "doctype": "Cost Center",
            "cost_center_name": "Test MT4000084238",
            "company": cls.company,
            "parent_cost_center": root_cc,
            "is_group": 0,
            "custom_machine_code": MACHINE_2_CODE,
        }).insert(ignore_permissions=True)

        frappe.db.commit()

        # --- BOM Upload: Machine 1 ---
        cls.result_1, cls.upload_name = run_bom_upload(
            cls.project.name, MACHINE_1_CODE, MACHINE_1_ROWS
        )

        # --- BOM Upload: Machine 2 (reuse same BOM Upload doc, different machine) ---
        cls.result_2, _ = run_bom_upload(
            cls.project.name, MACHINE_2_CODE, MACHINE_2_ROWS,
            upload_doc_name=cls.upload_name,
        )

    @classmethod
    def tearDownClass(cls):
        """Clean up all test data"""
        # MRs
        for mr in frappe.db.get_all("Material Request",
                                     {"custom_project_": cls.project.name}, pluck="name"):
            frappe.get_doc("Material Request", mr).cancel()
            frappe.delete_doc("Material Request", mr, force=True)

        # CMs
        for cm in frappe.db.get_all("Project Component Master",
                                     {"project": cls.project.name}, pluck="name"):
            frappe.delete_doc("Project Component Master", cm, force=True)

        # BOMs (cancel first, then delete)
        test_items = ["MT4000084237", "MT4000084238", "GT3000012345",
                      "GT3000067890", "GT3000099999", "DT0000054321",
                      "DT0000054322", "DT0000054323"]
        for item_code in test_items:
            for bom in frappe.db.get_all("BOM", {"item": item_code, "docstatus": 1}, pluck="name"):
                frappe.get_doc("BOM", bom).cancel()
            for bom in frappe.db.get_all("BOM", {"item": item_code}, pluck="name"):
                frappe.delete_doc("BOM", bom, force=True)

        # BOM Upload doc
        if frappe.db.exists("BOM Upload", cls.upload_name):
            frappe.delete_doc("BOM Upload", cls.upload_name, force=True)

        # Items
        all_items = test_items + [RAW_MATERIAL, RAW_MATERIAL_2]
        for item_code in all_items:
            if frappe.db.exists("Item", item_code):
                frappe.delete_doc("Item", item_code, force=True)

        # Cost Centers
        for cc in [cls.cost_center_1, cls.cost_center_2]:
            frappe.delete_doc("Cost Center", cc.name, force=True)

        # Project
        frappe.delete_doc("Project", cls.project.name, force=True)

        frappe.db.commit()
        super().tearDownClass()

    def setUp(self):
        frappe.set_user("Administrator")

    def tearDown(self):
        """Delete MRs created during this test"""
        for mr in frappe.db.get_all("Material Request",
                                     {"custom_project_": self.project.name}, pluck="name"):
            try:
                doc = frappe.get_doc("Material Request", mr)
                if doc.docstatus == 1:
                    doc.cancel()
                frappe.delete_doc("Material Request", mr, force=True)
            except Exception:
                pass
        frappe.db.commit()

    # --- Helpers ---

    def _make_mr(self, item_code, qty, machine_code, cost_center, bom_no=None):
        """Create a test Material Request"""
        item_row = {
            "item_code": item_code,
            "qty": qty,
            "schedule_date": frappe.utils.today(),
            "warehouse": "Material Staging - CT",
        }
        if bom_no:
            item_row["bom_no"] = bom_no

        return frappe.get_doc({
            "doctype": "Material Request",
            "material_request_type": "Purchase",
            "custom_project_": self.project.name,
            "custom_machine_code": machine_code,
            "custom_cost_center": cost_center,
            "items": [item_row],
        })

    def _get_bom_for_item(self, item_code):
        """Get the active default BOM for an item"""
        return frappe.db.get_value(
            "BOM", {"item": item_code, "is_active": 1, "is_default": 1, "docstatus": 1}, "name"
        )

    # =========================================================================
    # SETUP VERIFICATION — confirm upload created what we expect
    # =========================================================================

    def test_00_upload_succeeded(self):
        """Verify BOM upload completed successfully"""
        print("\n=== Machine 1 Upload Result ===")
        print(f"Status: {self.result_1.get('status')}")
        print(f"Summary: {self.result_1.get('summary')}")
        if self.result_1.get('errors'):
            print(f"Errors: {self.result_1.get('errors')}")

        print("\n=== Machine 2 Upload Result ===")
        print(f"Status: {self.result_2.get('status')}")
        print(f"Summary: {self.result_2.get('summary')}")
        if self.result_2.get('errors'):
            print(f"Errors: {self.result_2.get('errors')}")

        self.assertEqual(self.result_1.get("status"), "success",
                         f"Machine 1 upload failed: {self.result_1}")
        self.assertEqual(self.result_2.get("status"), "success",
                         f"Machine 2 upload failed: {self.result_2}")

    def test_01_d_code_defaults_to_buy(self):
        """D-codes should default to Buy after upload"""
        # Diagnostic: dump all CM values for the project
        all_cms = frappe.db.get_all(
            "Project Component Master",
            filters={"project": self.project.name},
            fields=["item_code", "machine_code", "make_or_buy", "project_qty", "bom_qty_required", "total_qty_limit"],
            order_by="machine_code, item_code"
        )
        print("\n=== ALL CM Values After Upload ===")
        for cm in all_cms:
            print(f"{cm.item_code} (machine={cm.machine_code}): make_or_buy={cm.make_or_buy}, "
                  f"project_qty={cm.project_qty}, bom_qty_required={cm.bom_qty_required}, "
                  f"total_qty_limit={cm.total_qty_limit}")

        cm = frappe.db.get_value(
            "Project Component Master",
            {"project": self.project.name, "item_code": D_CODE_1,
             "machine_code": MACHINE_1_CODE},
            ["make_or_buy", "total_qty_limit"],
            as_dict=True,
        )
        self.assertIsNotNone(cm, f"CM not found for {D_CODE_1}")
        self.assertEqual(cm.make_or_buy, "Buy")
        self.assertGreater(cm.total_qty_limit, 0,
                           "D-code (Buy) should have limit > 0")

    def test_02_rm_under_buy_parent_has_zero_limit(self):
        """RM under Buy D-code should have total_qty_limit = 0 (not needed)"""
        cm = frappe.db.get_value(
            "Project Component Master",
            {"project": self.project.name, "item_code": RAW_MATERIAL,
             "machine_code": MACHINE_1_CODE},
            ["total_qty_limit", "bom_qty_required", "make_or_buy"],
            as_dict=True,
        )
        self.assertIsNotNone(cm, f"CM not found for {RAW_MATERIAL}")
        self.assertEqual(cm.make_or_buy, "Buy")
        self.assertEqual(cm.total_qty_limit, 0,
                         "RM under Buy parent should have limit=0")

    def test_03_cross_machine_creates_separate_cm(self):
        """Same RM in 2 machines should have 2 separate CMs"""
        cm_count = frappe.db.count(
            "Project Component Master",
            {"project": self.project.name, "item_code": RAW_MATERIAL},
        )
        self.assertEqual(cm_count, 2, f"Expected 2 CMs for {RAW_MATERIAL}, got {cm_count}")

    def test_04_rm_directly_under_machine_has_limit(self):
        """RM directly under M-code (Make parent) should have limit > 0"""
        cm = frappe.db.get_value(
            "Project Component Master",
            {"project": self.project.name, "item_code": RAW_MATERIAL_2,
             "machine_code": MACHINE_1_CODE},
            ["total_qty_limit", "make_or_buy"],
            as_dict=True,
        )
        self.assertIsNotNone(cm, f"CM not found for {RAW_MATERIAL_2}")
        self.assertEqual(cm.make_or_buy, "Buy")
        self.assertGreater(cm.total_qty_limit, 0,
                           "RM directly under Make parent should have limit > 0")

    # =========================================================================
    # MAKE / BUY ENFORCEMENT
    # =========================================================================

    def test_10_make_item_blocked_from_mr(self):
        """G-code items are 'Make' — cannot be added to Material Request"""
        mr = self._make_mr(G_CODE_1, 1, MACHINE_1_CODE, self.cost_center_1.name)

        with self.assertRaises(frappe.ValidationError) as ctx:
            mr.insert()

        self.assertIn("Make", str(ctx.exception))

    def test_11_buy_d_code_allowed_in_mr(self):
        """D-code (Buy) can be procured via MR"""
        mr = self._make_mr(D_CODE_1, 1, MACHINE_1_CODE, self.cost_center_1.name)
        mr.insert()
        self.assertTrue(mr.name)

    def test_12_rm_under_buy_parent_blocked_by_zero_limit(self):
        """RM under Buy D-code blocked — total_qty_limit = 0"""
        mr = self._make_mr(RAW_MATERIAL, 1, MACHINE_1_CODE, self.cost_center_1.name)

        with self.assertRaises(frappe.ValidationError) as ctx:
            mr.insert()

        self.assertIn("0", str(ctx.exception))  # limit = 0

    # =========================================================================
    # MAKE/BUY CASCADE — Change D-codes to Make, then test G-code validation
    # =========================================================================

    def _change_d_codes_to_make(self):
        """Change all test D-codes to Make and trigger recalculation"""
        for d_code in [D_CODE_1, D_CODE_2, D_CODE_3]:
            cm_name = frappe.db.get_value(
                "Project Component Master",
                {"project": self.project.name, "item_code": d_code},
                "name",
            )
            if cm_name:
                cm = frappe.get_doc("Project Component Master", cm_name)
                cm.make_or_buy = "Make"
                cm.save(ignore_permissions=True)
        frappe.db.commit()
        # Clear cache to prevent stale document issues
        frappe.clear_cache()

    def _revert_d_codes_to_buy(self):
        """Revert D-codes back to Buy"""
        for d_code in [D_CODE_1, D_CODE_2, D_CODE_3]:
            cm_name = frappe.db.get_value(
                "Project Component Master",
                {"project": self.project.name, "item_code": d_code},
                "name",
            )
            if cm_name:
                cm = frappe.get_doc("Project Component Master", cm_name)
                cm.make_or_buy = "Buy"
                cm.save(ignore_permissions=True)
        frappe.db.commit()
        # Clear cache to prevent stale document issues
        frappe.clear_cache()

    def test_20_after_make_switch_rm_limit_recalculated(self):
        """After D-codes changed to Make, RM total_qty_limit should be 100 (60+40)"""
        self._change_d_codes_to_make()
        try:
            cm = frappe.db.get_value(
                "Project Component Master",
                {"project": self.project.name, "item_code": RAW_MATERIAL,
                 "machine_code": MACHINE_1_CODE},
                ["total_qty_limit", "bom_qty_required"],
                as_dict=True,
            )
            self.assertEqual(cm.total_qty_limit, 100,
                             f"Expected total_qty_limit=100, got {cm.total_qty_limit}")
        finally:
            self._revert_d_codes_to_buy()

    def test_21_after_make_switch_d_code_blocked_from_mr(self):
        """After D-code changed to Make, it should be blocked from MR"""
        self._change_d_codes_to_make()
        try:
            mr = self._make_mr(D_CODE_1, 1, MACHINE_1_CODE, self.cost_center_1.name)
            with self.assertRaises(frappe.ValidationError) as ctx:
                mr.insert()
            self.assertIn("Make", str(ctx.exception))
        finally:
            self._revert_d_codes_to_buy()

    # =========================================================================
    # G-CODE LEVEL VALIDATION (requires D-codes as Make so RM has limit > 0)
    # =========================================================================

    def test_30_mr_with_bom_no_validates_against_g_code_limit(self):
        """MR with bom_no should use G-code limit (60), not overall (100)"""
        self._change_d_codes_to_make()
        try:
            bom_g1 = self._get_bom_for_item(G_CODE_1)
            mr = self._make_mr(RAW_MATERIAL, 70, MACHINE_1_CODE,
                               self.cost_center_1.name, bom_no=bom_g1)

            with self.assertRaises(frappe.ValidationError) as ctx:
                mr.insert()

            self.assertIn(f"G-code {G_CODE_1}", str(ctx.exception))
        finally:
            self._revert_d_codes_to_buy()

    def test_31_mr_without_bom_no_validates_against_overall_limit(self):
        """MR without bom_no should use overall CM limit (100)"""
        self._change_d_codes_to_make()
        try:
            mr = self._make_mr(RAW_MATERIAL, 110, MACHINE_1_CODE,
                               self.cost_center_1.name)

            with self.assertRaises(frappe.ValidationError) as ctx:
                mr.insert()

            self.assertIn("overall limit", str(ctx.exception))
        finally:
            self._revert_d_codes_to_buy()

    def test_32_g_codes_have_independent_limits(self):
        """60 for G1 + 40 for G2 = 100 total. Both pass (each within own limit)"""
        self._change_d_codes_to_make()
        try:
            bom_g1 = self._get_bom_for_item(G_CODE_1)
            bom_g2 = self._get_bom_for_item(G_CODE_2)

            mr1 = self._make_mr(RAW_MATERIAL, 60, MACHINE_1_CODE,
                                self.cost_center_1.name, bom_no=bom_g1)
            mr1.insert()

            mr2 = self._make_mr(RAW_MATERIAL, 40, MACHINE_1_CODE,
                                self.cost_center_1.name, bom_no=bom_g2)
            mr2.insert()

            self.assertTrue(mr1.name)
            self.assertTrue(mr2.name)
        finally:
            self._revert_d_codes_to_buy()

    def test_33_cumulative_qty_across_multiple_mrs(self):
        """Second MR for same G-code counts first MR's qty (40 + 30 = 70 > 60)"""
        self._change_d_codes_to_make()
        try:
            bom_g1 = self._get_bom_for_item(G_CODE_1)

            mr1 = self._make_mr(RAW_MATERIAL, 40, MACHINE_1_CODE,
                                self.cost_center_1.name, bom_no=bom_g1)
            mr1.insert()

            mr2 = self._make_mr(RAW_MATERIAL, 30, MACHINE_1_CODE,
                                self.cost_center_1.name, bom_no=bom_g1)

            with self.assertRaises(frappe.ValidationError) as ctx:
                mr2.insert()

            self.assertIn(f"G-code {G_CODE_1}", str(ctx.exception))
        finally:
            self._revert_d_codes_to_buy()

    # =========================================================================
    # CROSS-MACHINE CM ISOLATION
    # =========================================================================

    def test_40_cross_machine_limits_are_independent(self):
        """Same RM on machine 2 should not share machine 1's limit"""
        self._change_d_codes_to_make()
        try:
            bom_g3 = self._get_bom_for_item(G_CODE_3)

            # Machine 2: 25 exceeds its limit of 20
            mr = self._make_mr(RAW_MATERIAL, 25, MACHINE_2_CODE,
                               self.cost_center_2.name, bom_no=bom_g3)

            with self.assertRaises(frappe.ValidationError) as ctx:
                mr.insert()

            self.assertIn("20", str(ctx.exception))
        finally:
            self._revert_d_codes_to_buy()

    def test_41_machine1_mr_does_not_affect_machine2(self):
        """MR on machine 1 should not reduce machine 2's available qty"""
        self._change_d_codes_to_make()
        try:
            bom_g1 = self._get_bom_for_item(G_CODE_1)
            bom_g3 = self._get_bom_for_item(G_CODE_3)

            # Use 60 on machine 1
            mr1 = self._make_mr(RAW_MATERIAL, 60, MACHINE_1_CODE,
                                self.cost_center_1.name, bom_no=bom_g1)
            mr1.insert()

            # Machine 2: 20 should still pass
            mr2 = self._make_mr(RAW_MATERIAL, 20, MACHINE_2_CODE,
                                self.cost_center_2.name, bom_no=bom_g3)
            mr2.insert()

            self.assertTrue(mr2.name)
        finally:
            self._revert_d_codes_to_buy()
