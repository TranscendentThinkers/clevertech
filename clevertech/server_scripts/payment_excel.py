import difflib
import re

import frappe
from openpyxl import load_workbook
from frappe.utils import getdate
import xlrd
import os


# =========================================================
# HELPERS
# =========================================================

def _normalize(name):
    return name.lower().replace(".", "").replace(",", "").strip()

def _parse_amount(val):
    """Parse numeric or string amounts like '11246.00 Cr' → 11246.0"""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    cleaned = re.sub(r'[^\d.]', '', str(val).replace(",", ""))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


# =========================================================
# FILE LOADER
# =========================================================

def _load_payment_rows(file_url):
    file_doc  = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()
    ext       = os.path.splitext(file_path)[1].lower()
    rows      = []

    if ext == ".xlsx":
        wb = load_workbook(file_path, data_only=True)
        rows = list(wb.active.iter_rows(values_only=True))
    elif ext == ".xls":
        book = xlrd.open_workbook(file_path)
        sheet = book.sheet_by_index(0)
        for r in range(sheet.nrows):
            rows.append(sheet.row_values(r))
    else:
        frappe.throw("❌ Unsupported file format. Upload .xls or .xlsx only.")

    return rows


# =========================================================
# COLUMN CONSTANTS
# Col 0: Date | Col 1: Particulars | Col 5: Vch No.
# Col 6: Debit Amount | Col 7: Credit Amount
# =========================================================

COL_DATE        = 0
COL_PARTICULARS = 1
COL_VCH_NO      = 5
COL_DEBIT       = 6
COL_CREDIT      = 7

EXCLUDE_PREFIX = ["agst ref", "against ref", "new ref", "reference", "ref", "bill ref"]
INCLUDE_KEYWORDS = [
    "being", "neft", "imps", "rtgs", "upi", "cms",
    "flight", "ticket", "pnr", "booking", "travel",
    "payment", "transfer", "air india", "indigo", "spicejet",
]


# =========================================================
# PARSER — splits rows into payment blocks
# =========================================================

def _parse_payment_blocks(rows):
    """
    Returns list of blocks:
    {
        date, party_name, party (ERP supplier name — exact match only),
        paid_from (ERP account name or None), paid_from_name (raw text),
        paid_amount, voucher_no, remarks,
        has_unresolved: bool   (True if supplier or paid_from missing)
    }
    """
    COMPANY        = frappe.defaults.get_user_default("Company")
    PAID_TO        = "Creditors - CT"
    MODE_OF_PAYMENT = "NEFT"

    blocks      = []
    current_doc = None

    def flush():
        if not current_doc:
            return
        has_unresolved = (
            not current_doc.get("party")
            or not current_doc.get("paid_from")
            or current_doc.get("paid_amount", 0) <= 0
        )
        blocks.append({**current_doc, "has_unresolved": has_unresolved,
                        "company": COMPANY, "paid_to": PAID_TO,
                        "mode_of_payment": MODE_OF_PAYMENT})

    for r in rows[7:]:   # data starts at row 8 (index 7)
        try:
            if len(r) < 8:
                continue

            date_val    = r[COL_DATE]
            particulars = r[COL_PARTICULARS]
            voucher_no  = r[COL_VCH_NO]
            debit       = _parse_amount(r[COL_DEBIT])
            credit      = _parse_amount(r[COL_CREDIT])
            text        = str(particulars).strip() if particulars else ""

            # ---- New payment block ----
            if date_val:
                flush()
                # Exact supplier lookup only — fuzzy is for display in validate
                supplier = frappe.db.get_value("Supplier", {"supplier_name": text}, "name")
                current_doc = {
                    "date":          getdate(date_val),
                    "party_name":    text,
                    "party":         supplier,
                    "paid_from":     None,
                    "paid_from_name": None,
                    "paid_amount":   0.0,
                    "voucher_no":    str(voucher_no).strip() if voucher_no else None,
                    "remarks":       "",
                }
                continue

            if not current_doc:
                continue

            # ---- Credit row → Bank/Cash account (Paid From) ----
            if credit and text:
                acc = frappe.db.get_value("Account", {"account_name": text}, "name")
                if acc:
                    current_doc["paid_from"]      = acc
                    current_doc["paid_from_name"] = text
                    current_doc["paid_amount"]    = credit
                else:
                    current_doc["paid_from_name"] = text   # unresolved, record raw name

            # ---- Debit row → fallback amount ----
            elif debit and current_doc["paid_amount"] == 0:
                current_doc["paid_amount"] = debit

            # ---- Remarks logic ----
            if text and not date_val and not debit and not credit:
                low = text.lower()
                if _normalize(text) == _normalize(current_doc.get("party_name", "")):
                    continue
                if frappe.db.get_value("Account", {"account_name": text}, "name"):
                    continue
                if any(low.startswith(p) for p in EXCLUDE_PREFIX):
                    continue
                if not any(k in low for k in INCLUDE_KEYWORDS):
                    continue
                if current_doc["remarks"]:
                    current_doc["remarks"] += " | "
                current_doc["remarks"] += text

        except Exception:
            pass

    flush()
    return blocks


# =========================================================
# API 1 — VALIDATE  (pre-flight, no PEs created)
# =========================================================

@frappe.whitelist()
def validate_payment_excel(file_url):
    """
    Returns:
    {
        total, matched, skipped,
        unresolved: [{name, issue_type, context, suggestions}]
    }
    """
    rows   = _load_payment_rows(file_url)
    blocks = _parse_payment_blocks(rows)

    matched = sum(1 for b in blocks if not b["has_unresolved"])
    skipped = len(blocks) - matched

    seen = set()
    unresolved_list = []

    for b in blocks:
        if not b["has_unresolved"]:
            continue
        vno = b.get("voucher_no") or b.get("party_name") or ""

        if not b["party"]:
            key = ("supplier", b["party_name"])
            if key not in seen:
                seen.add(key)
                unresolved_list.append({
                    "name":        b["party_name"],
                    "issue_type":  "Supplier not found",
                    "context":     vno,
                    "suggestions": [],
                })

        if not b["paid_from"] and b.get("paid_from_name"):
            key = ("account", b["paid_from_name"])
            if key not in seen:
                seen.add(key)
                unresolved_list.append({
                    "name":        b["paid_from_name"],
                    "issue_type":  "Bank/Cash account not found",
                    "context":     vno,
                    "suggestions": [],
                })

        if b["paid_amount"] <= 0:
            key = ("amount", vno)
            if key not in seen:
                seen.add(key)
                unresolved_list.append({
                    "name":        vno,
                    "issue_type":  "Paid amount is zero",
                    "context":     vno,
                    "suggestions": [],
                })

    # Fuzzy suggestions for unresolved supplier names
    missing_supplier_names = list({
        b["party_name"] for b in blocks if not b["party"]
    })
    if missing_supplier_names:
        all_suppliers = frappe.db.sql_list(
            "SELECT supplier_name FROM `tabSupplier` WHERE disabled = 0"
        )
        all_accounts = frappe.db.sql_list(
            "SELECT account_name FROM `tabAccount` WHERE disabled = 0"
        )
        all_names = all_suppliers + all_accounts
        sugg_map = {
            name: difflib.get_close_matches(name, all_names, n=3, cutoff=0.6)
            for name in missing_supplier_names
        }
        for u in unresolved_list:
            if u["issue_type"] == "Supplier not found":
                u["suggestions"] = sugg_map.get(u["name"], [])

    return {
        "total":      len(blocks),
        "matched":    matched,
        "skipped":    skipped,
        "unresolved": unresolved_list,
    }


# =========================================================
# API 2 — UPLOAD  (creates PEs for matched blocks only)
# =========================================================

@frappe.whitelist()
def upload_payment_excel(file_url):
    rows   = _load_payment_rows(file_url)
    blocks = _parse_payment_blocks(rows)

    created   = 0
    skipped   = 0
    duplicate = 0

    for b in blocks:
        if b["has_unresolved"]:
            skipped += 1
            continue
        vno = b.get("voucher_no")
        if vno and frappe.db.exists("Payment Entry", {"custom_tally_voucher_no": vno}):
            duplicate += 1
            continue
        try:
            _create_payment_entry(b)
            created += 1
        except Exception:
            skipped += 1

    parts = [f"✅ {created} created"]
    if skipped:
        parts.append(f"⚠️ {skipped} skipped (unresolved accounts/suppliers)")
    if duplicate:
        parts.append(f"🔁 {duplicate} duplicate (voucher no already exists)")

    return {"created": created, "skipped": skipped, "duplicate": duplicate, "message": " | ".join(parts)}


# =========================================================
# CREATOR
# =========================================================

def _create_payment_entry(b):
    pe = frappe.new_doc("Payment Entry")
    pe.payment_type             = "Pay"
    pe.posting_date             = b["date"]
    pe.company                  = b["company"]
    pe.mode_of_payment          = b["mode_of_payment"]
    pe.party_type               = "Supplier"
    pe.party                    = b["party"]
    pe.party_name               = b["party_name"]
    pe.paid_from                = b["paid_from"]
    pe.paid_to                  = b["paid_to"]
    pe.paid_amount              = b["paid_amount"]
    pe.received_amount          = b["paid_amount"]
    pe.reference_date           = b["date"]
    pe.custom_tally_voucher_no  = b["voucher_no"]
    pe.remarks                  = b["remarks"]
    pe.docstatus                = 0
    pe.insert(ignore_permissions=True)
