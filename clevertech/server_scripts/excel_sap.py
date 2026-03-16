import difflib

import frappe
from openpyxl import load_workbook
from frappe.utils import getdate
import xlrd
import os


# =========================================================
# FILE LOADER + HEADER/COLUMN DETECTION (shared)
# =========================================================

def _load_rows(file_url):
    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = file_doc.get_full_path()
    ext = os.path.splitext(file_path)[1].lower()
    rows = []

    if ext == ".xlsx":
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    elif ext == ".xls":
        book = xlrd.open_workbook(file_path)
        sheet = book.sheet_by_index(0)
        for r in range(sheet.nrows):
            rows.append(sheet.row_values(r))
    else:
        frappe.throw("❌ Unsupported file format. Upload .xls or .xlsx only.")

    def normalize(text):
        return str(text).lower().replace(" ", "").replace("_", "").strip()

    def find_header_row(rows):
        for idx, row in enumerate(rows):
            row_norm = [normalize(c) if c else "" for c in row]
            if (any("date" in c for c in row_norm)
                    and any("particular" in c for c in row_norm)
                    and any("debit" in c for c in row_norm)
                    and any("credit" in c for c in row_norm)):
                return idx
        return None

    header_row_index = find_header_row(rows)
    if header_row_index is None:
        frappe.throw("❌ Could not detect header row. Expected columns: Date, Particulars, Debit, Credit.")

    headers = [normalize(h) if h else "" for h in rows[header_row_index]]

    def find_col(keywords):
        for i, h in enumerate(headers):
            for k in keywords:
                if k in h:
                    return i
        return None

    cols = {
        "date":      find_col(["date"]),
        "particular": find_col(["particular"]),
        "debit":     find_col(["debit"]),
        "credit":    find_col(["credit"]),
        "voucher":   find_col(["vchno", "voucher", "documentno", "docno"]),
    }

    missing = []
    if cols["date"] is None: missing.append("Date")
    if cols["particular"] is None: missing.append("Particulars")
    if cols["debit"] is None: missing.append("Debit")
    if cols["credit"] is None: missing.append("Credit")
    if missing:
        frappe.throw(f"❌ Missing required columns in Excel: {', '.join(missing)}")

    return rows, cols, header_row_index + 1  # data_start_row


# =========================================================
# PARSER — splits rows into journal blocks
# =========================================================

def _parse_journal_blocks(rows, cols, data_start_row):
    """
    Returns a list of blocks, each:
    {
        header:         {posting_date, voucher_no, title, remark},
        resolved:       [{account, party_type, party, debit, credit, name}],
        unresolved:     [{name, suggested_type, debit, credit}],
        has_unresolved: bool,
        unbalanced:     bool,   # True only when resolved legs don't balance
        total_debit:    float,
        total_credit:   float,
    }
    Entries with unresolved legs are flagged so the caller can decide whether
    to skip or create them.
    """
    COL_DATE      = cols["date"]
    COL_PARTICULAR = cols["particular"]
    COL_DEBIT     = cols["debit"]
    COL_CREDIT    = cols["credit"]
    COL_VOUCHER   = cols["voucher"]

    def safe(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    blocks = []
    current_header = None
    resolved_buf   = []
    unresolved_buf = []

    def flush():
        nonlocal current_header, resolved_buf, unresolved_buf
        if not current_header:
            return
        # Use ALL rows (resolved + unresolved) to check if the entry truly balances.
        all_rows   = resolved_buf + unresolved_buf
        all_dr     = sum(r["debit"]  for r in all_rows)
        all_cr     = sum(r["credit"] for r in all_rows)
        # "Truly unbalanced" = every account name was found in ERP (nothing missing)
        # but the amounts still don't balance — a genuine data error.
        # Entries with unresolved legs are skipped for a different reason (missing accounts),
        # so we don't double-count them as unbalanced.
        truly_unbalanced = not bool(unresolved_buf) and round(all_dr, 2) != round(all_cr, 2)
        blocks.append({
            "header":         current_header,
            "resolved":       resolved_buf,
            "unresolved":     unresolved_buf,
            "has_unresolved": bool(unresolved_buf),
            "unbalanced":     truly_unbalanced,
            "total_debit":    all_dr,
            "total_credit":   all_cr,
        })
        current_header = None
        resolved_buf   = []
        unresolved_buf = []

    for r in rows[data_start_row:]:
        try:
            posting_date = safe(r, COL_DATE)
            particulars  = safe(r, COL_PARTICULAR)
            voucher_no   = safe(r, COL_VOUCHER)
            debit        = safe(r, COL_DEBIT)  or 0
            credit       = safe(r, COL_CREDIT) or 0
            text         = str(particulars).strip() if particulars else ""

            # ---- New journal entry starts when a date appears ----
            if posting_date:
                flush()
                current_header = {
                    "posting_date": getdate(posting_date),
                    "voucher_no":   str(voucher_no).strip() if voucher_no else None,
                    "title":        text,
                    "remark":       "",
                }
                # Fall through — the same row may also carry the first accounting leg

            if not current_header:
                continue

            # ---- Narration row (text but no amounts) ----
            if text and not debit and not credit:
                if current_header["remark"]:
                    current_header["remark"] += "\n"
                current_header["remark"] += text
                continue

            if not text or (not debit and not credit):
                continue

            # ---- Accounting row ----
            account_data = _resolve_account_or_supplier(text)
            if account_data:
                resolved_buf.append({
                    "account":    account_data["account"],
                    "party_type": account_data.get("party_type"),
                    "party":      account_data.get("party"),
                    "debit":      float(debit  or 0),
                    "credit":     float(credit or 0),
                    "name":       text,
                })
            else:
                lname = text.lower()
                suggested_type = (
                    "Supplier" if any(k in lname for k in ["ltd", "private", "pvt", "limited"])
                    else "Account"
                )
                unresolved_buf.append({
                    "name":           text,
                    "suggested_type": suggested_type,
                    "debit":          float(debit  or 0),
                    "credit":         float(credit or 0),
                })

        except Exception:
            pass  # skip malformed rows silently in parse pass

    flush()
    return blocks


# =========================================================
# ACCOUNT / SUPPLIER RESOLVER
# =========================================================

def _resolve_account_or_supplier(name):
    acc = frappe.db.get_value("Account", {"account_name": name}, "name")
    if acc:
        return {"account": acc}

    sup = frappe.db.get_value("Supplier", {"supplier_name": name}, "name")
    if sup:
        try:
            payable = _get_supplier_payable_account(sup)
            return {"account": payable, "party_type": "Supplier", "party": sup}
        except Exception:
            # Supplier exists in ERP but has no payable account configured.
            # Treat as unresolved so the row isn't silently dropped.
            return None

    return None


def _get_supplier_payable_account(supplier):
    company = frappe.defaults.get_user_default("Company")
    payable = frappe.db.get_value(
        "Party Account", {"party": supplier, "company": company}, "account"
    )
    if payable:
        return payable
    payable = frappe.db.get_value("Company", company, "default_payable_account")
    if payable:
        return payable
    frappe.throw(f"❌ No payable account found for Supplier '{supplier}' in company '{company}'")


# =========================================================
# API 1 — VALIDATE  (pre-flight, no JEs created)
# =========================================================

@frappe.whitelist()
def validate_sap_journal(file_url):
    """
    Parse the Excel file and return a validation summary WITHOUT creating any
    Journal Entries. The frontend uses this to show unmatched accounts/suppliers
    and let the user decide whether to proceed with the matched subset.

    Returns:
    {
        total:      int,   # total journal blocks found
        matched:    int,   # blocks where all accounts resolved and balanced
        skipped:    int,   # blocks that will be skipped on upload
        unresolved: [{name, suggested_type, voucher_no}],   # deduplicated
        unbalanced: [{voucher_no, debit, credit, diff}],
    }
    """
    rows, cols, data_start_row = _load_rows(file_url)
    blocks = _parse_journal_blocks(rows, cols, data_start_row)

    seen_unresolved = set()
    unresolved_list = []
    unbalanced_list = []

    for b in blocks:
        vno = b["header"].get("voucher_no") or b["header"].get("title") or ""
        for u in b["unresolved"]:
            if u["name"] not in seen_unresolved:
                seen_unresolved.add(u["name"])
                unresolved_list.append({
                    "name":           u["name"],
                    "suggested_type": u["suggested_type"],
                    "voucher_no":     vno,
                })
        if b["unbalanced"]:
            unbalanced_list.append({
                "voucher_no": vno,
                "debit":      b["total_debit"],
                "credit":     b["total_credit"],
                "diff":       round(b["total_debit"] - b["total_credit"], 2),
            })

    matched = sum(1 for b in blocks if not b["has_unresolved"] and not b["unbalanced"])
    skipped = len(blocks) - matched

    # ---- Fuzzy suggestions for unresolved names ----
    # Load all active account names + supplier names once, then match.
    all_accounts  = frappe.db.sql_list(
        "SELECT account_name FROM `tabAccount` WHERE disabled = 0"
    )
    all_suppliers = frappe.db.sql_list(
        "SELECT supplier_name FROM `tabSupplier` WHERE disabled = 0"
    )
    all_names = all_accounts + all_suppliers

    for u in unresolved_list:
        suggestions = difflib.get_close_matches(u["name"], all_names, n=3, cutoff=0.6)
        u["suggestions"] = suggestions

    return {
        "total":      len(blocks),
        "matched":    matched,
        "skipped":    skipped,
        "unresolved": unresolved_list,
        "unbalanced": unbalanced_list,
    }


# =========================================================
# API 2 — UPLOAD  (creates JEs for matched blocks only)
# =========================================================

@frappe.whitelist()
def upload_sap_journal(file_url):
    """
    Create Journal Entries for every block where ALL accounts resolved and
    the entry is balanced. Blocks with unresolved accounts or imbalances are
    skipped (they should have been shown to the user via validate_sap_journal).
    """
    rows, cols, data_start_row = _load_rows(file_url)
    blocks = _parse_journal_blocks(rows, cols, data_start_row)

    created   = 0
    skipped   = 0
    duplicate = 0
    for b in blocks:
        if b["has_unresolved"] or b["unbalanced"]:
            skipped += 1
            continue
        vno = b["header"].get("voucher_no")
        if vno and frappe.db.exists("Journal Entry", {"custom_tally_voucher_no": vno}):
            duplicate += 1
            continue
        try:
            _create_journal_entry(b["header"], b["resolved"])
            created += 1
        except Exception:
            skipped += 1

    parts = [f"✅ {created} created"]
    if skipped:
        parts.append(f"⚠️ {skipped} skipped (unresolved/unbalanced)")
    if duplicate:
        parts.append(f"🔁 {duplicate} duplicate (voucher no already exists)")

    return {
        "created":   created,
        "skipped":   skipped,
        "duplicate": duplicate,
        "message":   " | ".join(parts),
    }


# =========================================================
# JOURNAL ENTRY CREATOR
# =========================================================

def _create_journal_entry(header, rows):
    je = frappe.new_doc("Journal Entry")
    je.voucher_type = "Journal Entry"
    je.posting_date = header["posting_date"]
    je.company      = frappe.defaults.get_user_default("Company")
    je.user_remark  = header.get("remark") or header.get("title")
    je.custom_tally_voucher_no = header.get("voucher_no")
    je.docstatus = 0

    for r in rows:
        je.append("accounts", {
            "account":                    r["account"],
            "party_type":                 r.get("party_type"),
            "party":                      r.get("party"),
            "debit_in_account_currency":  r["debit"],
            "credit_in_account_currency": r["credit"],
        })

    je.insert(ignore_permissions=True)
