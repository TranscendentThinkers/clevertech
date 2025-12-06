import frappe
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from frappe.utils.file_manager import save_file
import io
@frappe.whitelist()
def upload_item_images_from_excel(file_url):
    import frappe
    import openpyxl
    from openpyxl_image_loader import SheetImageLoader
    from frappe.utils.file_manager import save_file
    import io

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    content = file_doc.get_content()

    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    image_loader = SheetImageLoader(ws)

    total_rows = ws.max_row - 1  # excluding header

    processed = 0
    missed = 0
    skipped = 0
    failed = 0
    logs = []
    missing_items = []

    for idx, row in enumerate(range(2, ws.max_row + 1), start=1):
        # realtime progress
        frappe.publish_progress(
            percent=(idx / total_rows) * 100 if total_rows else 100,
            title="Uploading Item Images",
            description=f"Processing row {idx} of {total_rows}"
        )

        item_code = ws[f"B{row}"].value
        image_cell = f"A{row}"
        item_description = ws[f"C{row}"].value

        try:
            if not image_loader.image_in(image_cell):
                missed += 1
                logs.append(f"Image Missing for:{item_code} at Row {idx+1}")
                continue

            if not frappe.db.exists("Item", item_code):
                failed += 1
                missing_items.append((item_code,item_description))
                logs.append(f"Item does not exists in system : {item_code} at Row {idx+1}")
                continue

            if frappe.db.get_value("Item", item_code, "image"):
                skipped += 1
                logs.append(f"Skipped (already has image): {item_code}")
                continue

            img = image_loader.get(image_cell)
            img_bytes = io.BytesIO()
            img.save(img_bytes, format="PNG")

            file_obj = save_file(
                f"{item_code}.png",
                img_bytes.getvalue(),
                "Item",
                item_code,
                is_private=0
            )

            frappe.db.set_value(
                "Item",
                item_code,
                "image",
                file_obj.file_url
            )

            processed += 1
            logs.append(f"Updated: {item_code}")

        except Exception as e:
            failed += 1
            logs.append(f"Error {item_code}: {str(e)}")
    missing_file_url = None
    if missing_items:
        from openpyxl import Workbook
        from frappe.utils.file_manager import save_file
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Missing Items"

        ws.append(["Item Code","Description"])


        for code,description in missing_items:
            ws.append([code,description])

        output = io.BytesIO()
        wb.save(output)

        missing_file = save_file(
            "missing_item_codes.xlsx",
            output.getvalue(),
            None,
            None,
            is_private=0
        )

        missing_file_url = missing_file.file_url
        frappe.log_error("Missing URL",missing_file.file_url)

    return {
        "processed": processed,
        "missed":missed,
        "skipped": skipped,
        "failed": failed,
        "logs": logs,
        "missing_file_url": missing_file_url
    }

